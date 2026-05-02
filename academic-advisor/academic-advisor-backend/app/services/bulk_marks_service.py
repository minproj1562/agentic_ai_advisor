# academic-advisor-backend/app/services/bulk_marks_service.py
""" Bulk Marks Upload Service — Mumbai University Marksheet Format
══════════════════════════════════════════════════════════════
Template columns:
  Sr.No | Seat No | Name | Subject(CA|MSE|ESE|TOT) | ... | Total | SGPI | Result
Mapping:           CA → internal_marks,  MSE+ESE → external_marks
"""

import io, csv, re, uuid, json, logging
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass, field as dc_field
import zipfile
from typing import Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter

from app.models.student_profile import (
    StudentProfile, SemesterRecord, SubjectScore,
)
from app.models.pending_marks import PendingStudentMarks
from app.core.curriculum import (
    get_semester_subjects, get_elective_options, SubjectDefinition,
)
from app.utils.password import generate_student_password, hash_password
# ── Added for university Excel parser ──
from app.services.university_excel_parser import (
    UniversityExcelParser, UniversityStudentResult,
)

logger = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════
# 1.  COMPONENT DEFINITIONS  (Mumbai University scheme)
# ══════════════════════════════════════════════════════════

def get_subject_components(sub: SubjectDefinition) -> List[Dict[str, Any]]:
    """
    Return the marking components for a subject in university format.

    Theory  → CA(20) + MSE(30) + ESE(50) = 100
    Lab     → IA(25) + PR(25)            =  50
    SBL     → TW(50) + PR(50)            = 100
    Project → TW(50)                     =  50
    Intern  → TW(100)                    = 100
    """
    ct = sub.course_type

    if ct in ("MNP", "MJP"):
        return [{"key": "TW", "label": "TW", "max": sub.internal_max}]

    if ct == "INT":
        return [{"key": "TW", "label": "TW", "max": sub.internal_max}]

    if ct == "SBL":
        return [
            {"key": "TW", "label": "TW", "max": sub.internal_max},
            {"key": "PR", "label": "PR", "max": sub.external_max},
        ]

    if ct == "LBC" or (sub.is_practical and ct not in ("SBL",)):
        return [
            {"key": "IA", "label": "IA", "max": sub.internal_max},
            {"key": "PR", "label": "PR/OR", "max": sub.external_max},
        ]

    # Theory  (PCC, BSC, ESC, AEC, PEC, OEC, etc.)
    ca = sub.internal_max  # 20
    if sub.external_max == 80:
        mse, ese = 30, 50
    elif sub.external_max == 0:
        return [{"key": "CA", "label": "CA", "max": ca}]
    else:
        mse = round(sub.external_max * 0.375)
        ese = sub.external_max - mse

    return [
        {"key": "CA", "label": "CA", "max": ca},
        {"key": "MSE", "label": "MSE", "max": mse},
        {"key": "ESE", "label": "ESE", "max": ese},
    ]


def components_to_subject_score(
    components: Dict[str, float],
    sub: SubjectDefinition,
) -> Tuple[float, float]:
    """
    Map component marks → (internal_marks, external_marks)

    Theory:  internal = CA,  external = MSE + ESE
    Lab:     internal = IA,  external = PR
    SBL:     internal = TW,  external = PR
    Project: internal = TW,  external = 0
    """
    ct = sub.course_type

    if ct in ("MNP", "MJP", "INT"):
        return components.get("TW", 0.0), 0.0

    if ct == "SBL":
        return components.get("TW", 0.0), components.get("PR", 0.0)

    if ct == "LBC" or sub.is_practical:
        return components.get("IA", 0.0), components.get("PR", 0.0)

    # Theory
    ca = components.get("CA", 0.0)
    mse = components.get("MSE", 0.0)
    ese = components.get("ESE", 0.0)
    return ca, mse + ese


def get_subject_type_label(sub: SubjectDefinition) -> str:
    labels = {
        "PCC": "Theory", "BSC": "Theory", "ESC": "Theory",
        "AEC": "Theory", "PEC": "Elective", "OEC": "Open Elec.",
        "LBC": "Lab", "SBL": "Skill Lab", "MNP": "Mini Proj",
        "MJP": "Major Proj", "INT": "Internship",
    }
    return labels.get(sub.course_type, sub.course_type)


# ══════════════════════════════════════════════════════════
# 2.  GRADE CALCULATION
# ══════════════════════════════════════════════════════════

def calculate_grade(total: float, max_marks: float = 100.0) -> Dict[str, Any]:
    if max_marks <= 0:
        max_marks = 100.0
    pct = (total / max_marks) * 100.0
    if pct >= 90: return {"grade": "O",  "points": 10.0}
    if pct >= 80: return {"grade": "A+", "points":  9.0}
    if pct >= 70: return {"grade": "A",  "points":  8.0}
    if pct >= 60: return {"grade": "B+", "points":  7.0}
    if pct >= 50: return {"grade": "B",  "points":  6.0}
    if pct >= 45: return {"grade": "C",  "points":  5.0}
    if pct >= 40: return {"grade": "P",  "points":  4.0}
    return {"grade": "F", "points": 0.0}


# ══════════════════════════════════════════════════════════
# 3.  DATA CLASSES
# ══════════════════════════════════════════════════════════

@dataclass
class ParsedStudentMarks:
    roll_number: str
    student_name: str
    subjects: List[Dict[str, Any]] = dc_field(default_factory=list)
    errors: List[str] = dc_field(default_factory=list)
    warnings: List[str] = dc_field(default_factory=list)
    sgpa_from_sheet: Optional[float] = None   # SGPI read directly from XLSX
    cgpa_from_sheet: Optional[float] = None   # CGPI read directly from XLSX


@dataclass
class UploadResult:
    upload_id: str = ""
    total_rows: int = 0
    matched_students: int = 0
    unmatched_students: int = 0
    updated_students: int = 0
    created_students: int = 0
    failed_updates: int = 0
    skipped_students: int = 0
    matched_details: List[Dict[str, Any]] = dc_field(default_factory=list)
    unmatched_roll_numbers: List[str] = dc_field(default_factory=list)
    errors: List[Dict[str, Any]] = dc_field(default_factory=list)
    warnings: List[str] = dc_field(default_factory=list)
    csv_data: str = ""
    semester: int = 0
    branch: str = ""
    academic_year: str = ""

    def to_dict(self) -> Dict[str, Any]:
        return {
            "upload_id": self.upload_id,
            "total_rows": self.total_rows,
            "matched_students": self.matched_students,
            "unmatched_students": self.unmatched_students,
            "updated_students": self.updated_students,
            "created_students": self.created_students,
            "failed_updates": self.failed_updates,
            "skipped_students": self.skipped_students,
            "matched_details": self.matched_details,
            "unmatched_roll_numbers": self.unmatched_roll_numbers,
            "errors": self.errors,
            "warnings": self.warnings,
            "csv_data": self.csv_data,
            "semester": self.semester,
            "branch": self.branch,
            "academic_year": self.academic_year,
        }


# ══════════════════════════════════════════════════════════
# 4.  STYLES
# ══════════════════════════════════════════════════════════

_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

STYLE = {
    "meta_font": Font(bold=True, size=11),
    "meta_fill": PatternFill("solid", fgColor="D6E4F0"),
    "hdr_font": Font(bold=True, color="FFFFFF", size=10),
    "hdr_fill": PatternFill("solid", fgColor="2F5496"),
    "comp_font": Font(bold=True, size=10),
    "comp_fill": PatternFill("solid", fgColor="B4C6E7"),
    "max_font": Font(italic=True, size=9, color="666666"),
    "max_fill": PatternFill("solid", fgColor="F2F2F2"),
    "elec_fill": PatternFill("solid", fgColor="E2EFDA"),
    "lab_fill": PatternFill("solid", fgColor="FCE4D6"),
    "proj_fill": PatternFill("solid", fgColor="F8CBAD"),
    "summary_fill": PatternFill("solid", fgColor="DDEBF7"),
    "inst_font": Font(italic=True, color="555555", size=10),
    "border": _BORDER,
}

CATEGORY_FILLS = {
    "PCC": PatternFill("solid", fgColor="2F5496"),
    "BSC": PatternFill("solid", fgColor="375623"),
    "ESC": PatternFill("solid", fgColor="548235"),
    "AEC": PatternFill("solid", fgColor="BF8F00"),
    "PEC": PatternFill("solid", fgColor="7030A0"),
    "OEC": PatternFill("solid", fgColor="00B0F0"),
    "LBC": PatternFill("solid", fgColor="C55A11"),
    "SBL": PatternFill("solid", fgColor="ED7D31"),
    "MNP": PatternFill("solid", fgColor="FF0066"),
    "MJP": PatternFill("solid", fgColor="CC0000"),
    "INT": PatternFill("solid", fgColor="002060"),
}


# ══════════════════════════════════════════════════════════
# 5.  SERVICE CLASS
# ══════════════════════════════════════════════════════════

class BulkMarksService:
    """
    Full pipeline:
      generate_template → admin fills → upload → parse → preview → save
    """

    # ─────────────────────────────────────────────────
    # 5.1  TEMPLATE GENERATION
    # ─────────────────────────────────────────────────

    def generate_template(
        self,
        semester: int,
        branch: str,
        academic_year: str,
        admission_year: int,
        elective_choices: Optional[Dict[str, str]] = None,
    ) -> io.BytesIO:
        """
        Generate a university-format XLSX template.

        Sheets:
          1. Marks Data   – the sheet admins fill in
          2. Subject Info  – read-only reference
          3. Elective Options – codes for each group (if any)
        """
        subjects = get_semester_subjects(semester, admission_year)
        if not subjects:
            raise ValueError(
                f"No subjects for semester {semester}, admission {admission_year}"
            )

        # Resolve elective choices
        resolved = self._resolve_electives(subjects, elective_choices)

        wb = Workbook()
        self._build_marks_sheet(wb, resolved, semester, branch, academic_year, admission_year)
        self._build_info_sheet(wb, resolved)

        elec = [s for s in subjects if s.is_elective and s.elective_group]
        if elec:
            self._build_elective_sheet(wb, elec)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def _resolve_electives(
        self,
        subjects: List[SubjectDefinition],
        choices: Optional[Dict[str, str]],
    ) -> List[SubjectDefinition]:
        """
        Replace elective placeholders with specific choices, or keep as-is.
        """
        if not choices:
            return subjects

        resolved = []
        for sub in subjects:
            if sub.is_elective and sub.elective_group and sub.elective_group in choices:
                code = choices[sub.elective_group]
                opts = get_elective_options(sub.elective_group)
                match = next((o for o in opts if o["code"].upper() == code.upper()), None)
                if match:
                    resolved.append(SubjectDefinition(
                        subject_code=match["code"],
                        subject_name=match["name"],
                        credits=sub.credits,
                        course_type=sub.course_type,
                        internal_max=sub.internal_max,
                        external_max=sub.external_max,
                        is_elective=True,
                        is_practical=sub.is_practical,
                        elective_group=sub.elective_group,
                    ))
                else:
                    resolved.append(sub)
            else:
                resolved.append(sub)
        return resolved

    # ── Marks Data sheet (university format) ──

    def _build_marks_sheet(self, wb, subjects, semester, branch, ay, adm_yr):
        ws = wb.active
        ws.title = "Marks Data"
        S = STYLE

        # ─── Row 1: metadata ───
        meta = [
            ("semester", semester), ("branch", branch),
            ("academic_year", ay), ("admission_year", adm_yr),
        ]
        c = 1
        for key, val in meta:
            k_cell = ws.cell(row=1, column=c, value=key)
            k_cell.font = S["meta_font"]; k_cell.fill = S["meta_fill"]
            v_cell = ws.cell(row=1, column=c + 1, value=val)
            v_cell.font = S["meta_font"]; v_cell.fill = S["meta_fill"]
            c += 2

        # ─── Row 2: instructions ───
        ws.cell(row=2, column=1, value=(
            "Instructions:  Fill marks below.  Seat No = Roll Number.  "
            "For electives, enter the chosen code in the ELEC CODE column.  "
            "See 'Subject Info' and 'Elective Options' sheets for reference.  "
            "TOT columns are auto-checked during upload (no formula needed)."
        )).font = S["inst_font"]
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=14)

        # ─── Row 3: empty spacer ───

        HDR = 4      # subject name row (merged)
        COMP = 5     # component row (CA, MSE, ESE, TOT)
        MAX_R = 6    # max-marks row
        DATA = 7     # first data row

        # ─── Fixed columns ───
        fixed_headers = [
            ("Sr.\nNo", 6),
            ("Seat No\n(Roll Number)", 20),
            ("Name of\nStudent", 30),
        ]
        for i, (label, width) in enumerate(fixed_headers, 1):
            cell = ws.cell(row=HDR, column=i, value=label)
            cell.font = S["hdr_font"]; cell.fill = S["hdr_fill"]
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = S["border"]
            ws.merge_cells(start_row=HDR, start_column=i, end_row=COMP, end_column=i)
            # re-apply styles after merge
            for r in (HDR, COMP):
                c2 = ws.cell(row=r, column=i)
                c2.border = S["border"]
            ws.column_dimensions[get_column_letter(i)].width = width

        # ─── Subject columns ───
        col = len(fixed_headers) + 1
        # Build column map for parser reference  (stored in hidden JSON row later)
        column_map: List[Dict[str, Any]] = []

        for sub in subjects:
            comps = get_subject_components(sub)
            n_comp = len(comps)
            n_cols = n_comp + 1  # +1 for TOT

            has_elective_code = sub.is_elective and sub.elective_group
            if has_elective_code:
                n_cols += 1  # extra column for elective code

            start_col = col

            # ── 4.1  merged subject header (row HDR) ──
            type_label = get_subject_type_label(sub)
            if has_elective_code:
                header_text = (
                    f"{sub.elective_group}\n({sub.subject_code}) "
                    f"[{sub.credits} cr, {type_label}]"
                )
            else:
                header_text = (
                    f"{sub.subject_name}\n({sub.subject_code}) "
                    f"[{sub.credits} cr, {type_label}]"
                )

            cell = ws.cell(row=HDR, column=start_col, value=header_text)
            fill = CATEGORY_FILLS.get(sub.course_type, S["hdr_fill"])
            cell.font = S["hdr_font"]; cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = S["border"]

            if n_cols > 1:
                ws.merge_cells(
                    start_row=HDR, start_column=start_col,
                    end_row=HDR, end_column=start_col + n_cols - 1,
                )

            # ── 4.2  component headers (row COMP) + max marks (row MAX_R) ──
            ci = start_col

            if has_elective_code:
                # Elective code column
                cc = ws.cell(row=COMP, column=ci, value="ELEC\nCODE")
                cc.font = S["comp_font"]; cc.fill = S["elec_fill"]
                cc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cc.border = S["border"]
                mc = ws.cell(row=MAX_R, column=ci, value="Code")
                mc.font = S["max_font"]; mc.fill = S["max_fill"]
                mc.alignment = Alignment(horizontal="center")
                mc.border = S["border"]
                ws.column_dimensions[get_column_letter(ci)].width = 14
                column_map.append({
                    "col": ci, "subject_code": sub.subject_code,
                    "component": "ELEC_CODE", "max": 0,
                    "elective_group": sub.elective_group,
                })
                ci += 1

            for comp in comps:
                cc = ws.cell(row=COMP, column=ci, value=comp["label"])
                cc.font = S["comp_font"]; cc.fill = S["comp_fill"]
                cc.alignment = Alignment(horizontal="center", vertical="center")
                cc.border = S["border"]

                mc = ws.cell(row=MAX_R, column=ci, value=comp["max"])
                mc.font = S["max_font"]; mc.fill = S["max_fill"]
                mc.alignment = Alignment(horizontal="center")
                mc.border = S["border"]

                ws.column_dimensions[get_column_letter(ci)].width = 8
                column_map.append({
                    "col": ci, "subject_code": sub.subject_code,
                    "component": comp["key"], "max": comp["max"],
                    "elective_group": sub.elective_group or "",
                })
                ci += 1

            # TOT column
            tc = ws.cell(row=COMP, column=ci, value="TOT")
            tc.font = Font(bold=True, size=10); tc.fill = S["comp_fill"]
            tc.alignment = Alignment(horizontal="center")
            tc.border = S["border"]
            total_max = sum(comp["max"] for comp in comps)
            tm = ws.cell(row=MAX_R, column=ci, value=total_max)
            tm.font = S["max_font"]; tm.fill = S["max_fill"]
            tm.alignment = Alignment(horizontal="center")
            tm.border = S["border"]
            ws.column_dimensions[get_column_letter(ci)].width = 8
            column_map.append({
                "col": ci, "subject_code": sub.subject_code,
                "component": "TOT", "max": total_max,
                "elective_group": sub.elective_group or "",
            })
            ci += 1

            col = ci

        # ── Summary columns ──
        summary_headers = ["Total\nMarks", "SGPI", "Result"]
        for label in summary_headers:
            cell = ws.cell(row=HDR, column=col, value=label)
            cell.font = S["hdr_font"]; cell.fill = S["summary_fill"]
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = S["border"]
            ws.merge_cells(start_row=HDR, start_column=col, end_row=COMP, end_column=col)
            ws.column_dimensions[get_column_letter(col)].width = 10
            col += 1

        # ── Max row styling for fixed cols ──
        for i in range(1, len(fixed_headers) + 1):
            c = ws.cell(row=MAX_R, column=i, value="")
            c.fill = S["max_fill"]; c.border = S["border"]

        # ── Empty data rows (50 students) ──
        total_cols = col - 1
        for r in range(DATA, DATA + 50):
            ws.cell(row=r, column=1, value=r - DATA + 1).alignment = Alignment(horizontal="center")
            for c2 in range(1, total_cols + 1):
                ws.cell(row=r, column=c2).border = S["border"]
                ws.cell(row=r, column=c2).alignment = Alignment(horizontal="center")

        # ── Store column map as JSON in a hidden row (row 100) for parser ──
        ws.cell(row=100, column=1, value="__COLUMN_MAP__")
        ws.cell(row=100, column=2, value=json.dumps(column_map))
        ws.row_dimensions[100].hidden = True

        # ── Freeze panes ──
        ws.freeze_panes = f"D{DATA}"

    # ── Subject Info sheet ──

    def _build_info_sheet(self, wb, subjects):
        ws = wb.create_sheet("Subject Info")
        S = STYLE
        headers = [
            "Code", "Name", "Credits", "Type", "Category",
            "Components", "Internal Max", "External Max", "Total Max",
            "Elective Group",
        ]
        hfont = Font(bold=True, color="FFFFFF")
        hfill = PatternFill("solid", fgColor="548235")

        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = hfont; c.fill = hfill; c.border = S["border"]

        for ri, sub in enumerate(subjects, 2):
            comps = get_subject_components(sub)
            comp_str = " + ".join(f"{c['label']}({c['max']})" for c in comps)
            vals = [
                sub.subject_code, sub.subject_name, sub.credits,
                sub.course_type, get_subject_type_label(sub),
                comp_str, sub.internal_max, sub.external_max,
                sub.internal_max + sub.external_max,
                sub.elective_group or "",
            ]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=ri, column=ci, value=v)
                c.border = S["border"]

        for ci in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 22

    # ── Elective Options sheet ──

    def _build_elective_sheet(self, wb, elective_subs):
        ws = wb.create_sheet("Elective Options")
        S = STYLE
        hfont = Font(bold=True, color="FFFFFF")
        hfill = PatternFill("solid", fgColor="BF8F00")

        for i, h in enumerate(["Elective Group", "Option Code", "Option Name"], 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = hfont; c.fill = hfill; c.border = S["border"]

        row = 2
        seen = set()
        for sub in elective_subs:
            grp = sub.elective_group
            if grp and grp not in seen:
                seen.add(grp)
                for opt in get_elective_options(grp):
                    ws.cell(row=row, column=1, value=grp).border = S["border"]
                    ws.cell(row=row, column=2, value=opt["code"]).border = S["border"]
                    ws.cell(row=row, column=3, value=opt["name"]).border = S["border"]
                    row += 1

        for ci in range(1, 4):
            ws.column_dimensions[get_column_letter(ci)].width = 30

    # ══════════════════════════════════════════════════════════════════
    # UNIVERSITY MARKSHEET PARSING (strict semester‑aware)
    # ══════════════════════════════════════════════════════════════════

    def parse_university_excel(
        self,
        file_bytes: bytes,
        semester: int,
        branch: str,
        academic_year: str,
        admission_year: int,
        sheet_name: Optional[str] = None,
        sheet_index: int = 0,
    ) -> Tuple[List[ParsedStudentMarks], Dict[str, Any]]:
        """
        Parse a Mumbai University marksheet Excel file.
        STRICT semester‑aware matching – only subjects belonging to the given semester are accepted.
        """
        from app.services.university_excel_parser import UniversityExcelParser

        # Get subjects for the target semester only
        curriculum_subjects = get_semester_subjects(semester, admission_year)
        if not curriculum_subjects:
            raise ValueError(
                f"No subjects in curriculum for semester {semester}, admission {admission_year}"
            )

        # Build strict lookup maps (exact, case‑insensitive)
        sub_map_by_code: Dict[str, SubjectDefinition] = {}
        sub_map_by_name: Dict[str, SubjectDefinition] = {}
        for sub in curriculum_subjects:
            sub_map_by_code[sub.subject_code.upper()] = sub
            sub_map_by_name[sub.subject_name.strip().lower()] = sub

        parser = UniversityExcelParser(file_bytes)
        try:
            uni_results, meta = parser.parse_single_sheet(
                sheet_name=sheet_name,
                sheet_index=sheet_index,
            )
        except ValueError as e:
            raise ValueError(f"University Excel parse error: {e}")

        parsed_list: List[ParsedStudentMarks] = []

        for uni_stu in uni_results:
            stu = ParsedStudentMarks(
                roll_number=uni_stu.seat_number,
                student_name=uni_stu.student_name,
                sgpa_from_sheet=uni_stu.sgpa,
                cgpa_from_sheet=uni_stu.cgpa,
            )

            for raw_subj in uni_stu.subjects:
                raw_code = raw_subj.get("subject_code", "").upper().strip()
                raw_name = raw_subj.get("subject_name", "").strip()

                # 1. Strict exact code match
                matched_sub = sub_map_by_code.get(raw_code)

                # 2. If code not found, try exact name match (case‑insensitive)
                if matched_sub is None and raw_name:
                    matched_sub = sub_map_by_name.get(raw_name.lower())

                # 3. If still no match, skip this subject entirely (do not guess)
                if matched_sub is None:
                    stu.warnings.append(
                        f"Subject '{raw_code}: {raw_name}' not found in curriculum for "
                        f"semester {semester}, admission {admission_year} – SKIPPED"
                    )
                    continue

                # 5. Use curriculum data (max marks, credits, etc.)
                internal = raw_subj.get("internal_marks", 0.0)
                external = raw_subj.get("external_marks", 0.0)
                total = raw_subj.get("total_marks", internal + external)
                max_tot = matched_sub.internal_max + matched_sub.external_max
                gi = calculate_grade(total, max_tot)

                # Validate marks against curriculum maxima
                if internal > matched_sub.internal_max + 0.5:
                    stu.warnings.append(
                        f"{raw_code}: internal {internal} > max {matched_sub.internal_max}"
                    )
                if external > matched_sub.external_max + 0.5:
                    stu.warnings.append(
                        f"{raw_code}: external {external} > max {matched_sub.external_max}"
                    )

                stu.subjects.append({
                    "subject_code": matched_sub.subject_code,
                    "subject_name": matched_sub.subject_name,
                    "credits": matched_sub.credits,
                    "internal_marks": round(internal, 2),
                    "external_marks": round(external, 2),
                    "internal_max": matched_sub.internal_max,
                    "external_max": matched_sub.external_max,
                    "total_marks": round(total, 2),
                    "grade": gi["grade"],
                    "grade_points": gi["points"],
                    "is_elective": matched_sub.is_elective,
                    "is_practical": matched_sub.is_practical,
                    "components": raw_subj.get("components", {}),
                })

            if not stu.subjects:
                stu.errors.append("No subjects could be matched to curriculum for this semester")
            else:
                logger.info(
                    f"Parsed {len(stu.subjects)} subjects for {stu.roll_number} (semester {semester})"
                )

            parsed_list.append(stu)

        meta.update({
            "format_detected": "university_marksheet",
            "semester": semester,
            "branch": branch,
            "academic_year": academic_year,
            "admission_year": admission_year,
            "subjects_in_curriculum": len(curriculum_subjects),
            "subjects_matched": sum(len(s.subjects) for s in parsed_list),
        })

        return parsed_list, meta

    def _detect_university_format(self, file_bytes: bytes, filename: str) -> bool:
        """
        Detect if an Excel file is a university marksheet format.
        Looks for characteristic markers: 'MarksO', 'C*GP', seat number patterns.
        """
        try:
            df = pd.read_excel(
                io.BytesIO(file_bytes),
                sheet_name=0,
                header=None,
                dtype=object,
                nrows=15,  # only peek at first 15 rows
            )

            # Check for 'MarksO' or 'C*GP' in any of the first 15 rows
            for _, row in df.iterrows():
                row_str = " ".join(str(v) for v in row.values if pd.notna(v))
                if "MarksO" in row_str or "C*GP" in row_str:
                    return True

            # Check for 'Subject Name' in column 3 area
            for i in range(min(15, len(df))):
                for c in range(min(6, df.shape[1])):
                    cell = df.iloc[i, c]
                    if pd.notna(cell) and str(cell).strip() in ("MarksO", "C*GP", "Subject Name"):
                        return True

            return False

        except Exception:
            return False

    # ─────────────────────────────────────────────────
    # 5.2  FILE PARSING (UPDATED)
    # ─────────────────────────────────────────────────

    def parse_file(
        self,
        file_bytes: bytes,
        filename: str,
        semester: int,
        branch: str,
        academic_year: str,
        admission_year: int,
    ) -> Tuple[List[ParsedStudentMarks], Dict[str, Any]]:
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

        # NEW: If it's a university marksheet (multi-row), process all sheets automatically
        if ext in ("xlsx", "xls") and self._detect_university_format(file_bytes, filename):
            return self.parse_multi_sheet_university(file_bytes, branch, academic_year, admission_year)

        # Legacy single-sheet handling for templates, CSV, etc.
        subjects = get_semester_subjects(semester, admission_year)
        sub_map = {s.subject_code: s for s in subjects}
        elec_map: Dict[str, Dict] = {}
        for s in subjects:
            if s.is_elective and s.elective_group:
                for opt in get_elective_options(s.elective_group):
                    elec_map[opt["code"].upper()] = {
                        "name": opt["name"],
                        "group": s.elective_group,
                        "template": s,
                    }

        meta = {
            "semester": semester, "branch": branch,
            "academic_year": academic_year, "admission_year": admission_year,
            "subjects_in_curriculum": len(subjects), "filename": filename,
        }

        if ext == "xlsx":
            parsed, fmt = self._parse_xlsx_university(file_bytes, sub_map, elec_map)
            meta["format_detected"] = fmt
            meta["total_rows"] = len(parsed)
            return parsed, meta

        if ext == "xls":
            parsed, fmt = self._parse_xls_fallback(file_bytes, sub_map, elec_map)
            meta["format_detected"] = fmt
            meta["total_rows"] = len(parsed)
            return parsed, meta

        if ext == "csv":
            parsed, fmt = self._parse_csv(file_bytes, sub_map, elec_map)
            meta["format_detected"] = fmt
            meta["total_rows"] = len(parsed)
            return parsed, meta

        raise ValueError(f"Unsupported file type .{ext}")

    def parse_multi_sheet_university(
        self,
        file_bytes: bytes,
        branch: str,
        academic_year: str,
        admission_year: int,
    ) -> Tuple[List[ParsedStudentMarks], Dict[str, Any]]:
        """
        Parse all sheets of a university marksheet Excel file.
        Automatically extracts semester from each sheet name.
        Returns combined parsed marks for all sheets.
        """
        from app.services.university_excel_parser import UniversityExcelParser, extract_semester_from_sheet_name

        parser = UniversityExcelParser(file_bytes)
        all_sheets_data = parser.parse_all_sheets()
        all_parsed = []
        combined_meta = {
            "sheets_processed": [],
            "total_students": 0,
            "total_subjects_matched": 0,
        }

        for sheet_name, sheet_info in all_sheets_data["sheets"].items():
            semester = extract_semester_from_sheet_name(sheet_name)
            if semester is None:
                logger.warning(f"Could not extract semester from sheet name '{sheet_name}' – skipping")
                continue

            curriculum_subjects = get_semester_subjects(semester, admission_year)
            if not curriculum_subjects:
                logger.warning(f"No curriculum subjects for semester {semester} – skipping")
                continue

            sub_map_by_code = {sub.subject_code.upper(): sub for sub in curriculum_subjects}
            sub_map_by_name = {sub.subject_name.strip().lower(): sub for sub in curriculum_subjects}

            sheet_parsed: List[ParsedStudentMarks] = []
            for uni_stu in sheet_info.get("results", []):
                stu = ParsedStudentMarks(
                    roll_number=uni_stu["seat_number"],
                    student_name=uni_stu["student_name"],
                    sgpa_from_sheet=uni_stu.get("sgpa"),
                    cgpa_from_sheet=uni_stu.get("cgpa"),
                )
                for raw_subj in uni_stu["subjects"]:
                    raw_code = raw_subj.get("subject_code", "").upper().strip()
                    raw_name = raw_subj.get("subject_name", "").strip()

                    matched_sub = sub_map_by_code.get(raw_code)
                    if matched_sub is None and raw_name:
                        matched_sub = sub_map_by_name.get(raw_name.lower())

                    if matched_sub is None:
                        stu.warnings.append(f"Subject '{raw_code}: {raw_name}' not found in semester {semester} – SKIPPED")
                        continue

                    if matched_sub.semester != semester:
                        stu.warnings.append(
                            f"Subject '{matched_sub.subject_code}' belongs to semester {matched_sub.semester}, "
                            f"but sheet '{sheet_name}' is semester {semester} – SKIPPED"
                        )
                        continue

                    internal = raw_subj.get("internal_marks", 0.0)
                    external = raw_subj.get("external_marks", 0.0)
                    total = raw_subj.get("total_marks", internal + external)
                    max_tot = matched_sub.internal_max + matched_sub.external_max
                    gi = calculate_grade(total, max_tot)

                    stu.subjects.append({
                        "subject_code": matched_sub.subject_code,
                        "subject_name": matched_sub.subject_name,
                        "credits": matched_sub.credits,
                        "internal_marks": round(internal, 2),
                        "external_marks": round(external, 2),
                        "internal_max": matched_sub.internal_max,
                        "external_max": matched_sub.external_max,
                        "total_marks": round(total, 2),
                        "grade": gi["grade"],
                        "grade_points": gi["points"],
                        "is_elective": matched_sub.is_elective,
                        "is_practical": matched_sub.is_practical,
                        "components": raw_subj.get("components", {}),
                    })

                if stu.subjects:
                    sheet_parsed.append(stu)
                else:
                    logger.debug(f"No subjects matched for student {uni_stu['seat_number']} in sheet '{sheet_name}'")

            combined_meta["sheets_processed"].append({
                "sheet_name": sheet_name,
                "detected_semester": semester,
                "students_parsed": len(sheet_parsed),
                "subjects_matched": sum(len(s.subjects) for s in sheet_parsed),
            })
            all_parsed.extend(sheet_parsed)

        combined_meta["total_students"] = len(all_parsed)
        combined_meta["total_subjects_matched"] = sum(len(s.subjects) for s in all_parsed)
        combined_meta["format_detected"] = "university_marksheet_multi_sheet"
        return all_parsed, combined_meta

    # ── XLSX parser (primary — handles our template with column map) ──

    def _parse_xlsx_university(
        self, content: bytes, sub_map, elec_map,
    ) -> Tuple[List[ParsedStudentMarks], str]:

        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active

        # ── Try reading embedded column map ──
        col_map_json = None
        try:
            marker = ws.cell(row=100, column=1).value
            if marker == "__COLUMN_MAP__":
                col_map_json = ws.cell(row=100, column=2).value
        except Exception:
            pass

        if col_map_json:
            return self._parse_with_column_map(
                ws, json.loads(col_map_json), sub_map, elec_map
            ), "university_template"

        # ── Fallback: auto-detect component row ──
        return self._parse_auto_detect(ws, sub_map, elec_map), "university_auto"

    def _parse_with_column_map(
        self, ws, col_map: List[Dict], sub_map, elec_map,
    ) -> List[ParsedStudentMarks]:
        """Parse using the embedded column map from our template."""

        # Find data start row (first non-empty row after max-marks row)
        data_start = 7  # default
        for r in range(4, 20):
            v = ws.cell(row=r, column=1).value
            if v is not None and str(v).strip().isdigit():
                data_start = r
                break

        # Group column map by subject
        subject_columns: Dict[str, Dict[str, Any]] = {}
        for entry in col_map:
            scode = entry["subject_code"]
            if scode not in subject_columns:
                subject_columns[scode] = {
                    "components": {},
                    "elective_group": entry.get("elective_group", ""),
                    "elec_code_col": None,
                }
            comp = entry["component"]
            if comp == "ELEC_CODE":
                subject_columns[scode]["elec_code_col"] = entry["col"]
            elif comp != "TOT":
                subject_columns[scode]["components"][comp] = {
                    "col": entry["col"], "max": entry["max"],
                }

        results: List[ParsedStudentMarks] = []

        for r in range(data_start, ws.max_row + 1):
            # Skip hidden metadata rows
            marker = ws.cell(row=r, column=1).value
            if marker is not None and str(marker).strip() == "__COLUMN_MAP__":
                continue
            if marker is not None and str(marker).strip() == "__STUDENTS_META__":
                continue

            roll_val = ws.cell(row=r, column=2).value  # Seat No column
            if roll_val is None or str(roll_val).strip() in ("", "nan", "None"):
                continue

            roll = str(roll_val).strip()

            # Skip if roll number looks like JSON
            if roll.startswith("[") or roll.startswith("{"):
                continue

            name = str(ws.cell(row=r, column=3).value or "").strip()
            stu = ParsedStudentMarks(roll_number=roll, student_name=name)

            for scode, info in subject_columns.items():
                # Resolve elective code if applicable
                act_code = scode
                act_name = ""
                tpl = sub_map.get(scode)
                is_elec = False

                if info["elec_code_col"]:
                    ec_val = ws.cell(row=r, column=info["elec_code_col"]).value
                    if ec_val:
                        ec = str(ec_val).strip().upper()
                        if ec in elec_map:
                            einfo = elec_map[ec]
                            act_code = ec
                            act_name = einfo["name"]
                            tpl = einfo["template"]
                            is_elec = True
                        else:
                            stu.warnings.append(f"Unknown elective code '{ec}'")
                            act_code = ec
                            tpl = sub_map.get(scode)
                            is_elec = True
                    else:
                        stu.warnings.append(
                            f"No elective code for {info['elective_group']} — skipped"
                        )
                        continue

                if tpl is None:
                    stu.warnings.append(f"Subject {scode} not in curriculum")
                    continue

                if not is_elec:
                    act_name = tpl.subject_name
                    is_elec = tpl.is_elective

                # Read component marks
                comp_marks: Dict[str, float] = {}
                for comp_key, comp_info in info["components"].items():
                    val = ws.cell(row=r, column=comp_info["col"]).value
                    mark = 0.0
                    if val is not None:
                        try:
                            mark = float(val)
                        except (ValueError, TypeError):
                            stu.errors.append(
                                f"{act_code}: invalid {comp_key} value '{val}'"
                            )
                    comp_marks[comp_key] = mark

                    # Validate
                    if mark < 0:
                        stu.errors.append(f"{act_code}: negative {comp_key}")
                    if mark > comp_info["max"]:
                        stu.errors.append(
                            f"{act_code}: {comp_key}={mark} > max {comp_info['max']}"
                        )

                # Map components → internal/external
                internal, external = components_to_subject_score(comp_marks, tpl)
                total = internal + external
                max_total = tpl.internal_max + tpl.external_max
                gi = calculate_grade(total, max_total)

                stu.subjects.append({
                    "subject_code": act_code,
                    "subject_name": act_name,
                    "credits": tpl.credits,
                    "internal_marks": internal,
                    "external_marks": external,
                    "internal_max": tpl.internal_max,
                    "external_max": tpl.external_max,
                    "total_marks": total,
                    "grade": gi["grade"],
                    "grade_points": gi["points"],
                    "is_elective": is_elec,
                    "is_practical": tpl.is_practical,
                    "components": comp_marks,
                })

            if not stu.subjects:
                stu.errors.append("No valid subjects found")
            results.append(stu)

        return results

    # ── Auto-detect parser (for non-template university sheets) ──

    def _parse_auto_detect(
        self, ws, sub_map, elec_map,
    ) -> List[ParsedStudentMarks]:
        """
        Auto-detect subject columns by looking for component patterns
        (CA, MSE, ESE, IA, PR, TW) in the worksheet rows.
        """
        # Find the component row
        comp_row = None
        for r in range(1, 15):
            vals = []
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if v:
                    vals.append(str(v).strip().upper())
            if any(v in ("CA", "MSE", "ESE", "IA", "PR", "TW") for v in vals):
                comp_row = r
                break

        if comp_row is None:
            raise ValueError(
                "Cannot detect university format — no CA/MSE/ESE/IA/PR headers found"
            )

        hdr_row = comp_row - 1  # subject names in the row above
        max_row = comp_row + 1  # max marks below

        # Map subject columns
        seat_col = roll_col = name_col = None
        subject_blocks: List[Dict[str, Any]] = []
        current_subject = None
        current_start = None

        for c in range(1, ws.max_column + 1):
            hdr_val = ws.cell(row=hdr_row, column=c).value
            comp_val = str(ws.cell(row=comp_row, column=c).value or "").strip().upper()

            if hdr_val:
                hdr_str = str(hdr_val).strip().lower()
                if "seat" in hdr_str or "roll" in hdr_str or "enrollment" in hdr_str:
                    seat_col = c; continue
                if hdr_str in ("name", "name of student", "student name", "student_name"):
                    name_col = c; continue
                if "sr" in hdr_str and "no" in hdr_str:
                    continue

                # Start of a new subject block
                if current_subject:
                    subject_blocks.append(current_subject)
                current_subject = {
                    "header": str(hdr_val).strip(),
                    "start_col": c,
                    "components": {},
                    "tot_col": None,
                }

            if comp_val in ("CA", "MSE", "ESE", "IA", "PR", "TW", "OR", "PR/OR"):
                comp_key = comp_val.replace("/OR", "").replace("PR/OR", "PR")
                if comp_key == "OR":
                    comp_key = "PR"
                max_val = ws.cell(row=max_row, column=c).value
                max_marks = float(max_val) if max_val else 0

                if current_subject:
                    current_subject["components"][comp_key] = {
                        "col": c, "max": max_marks,
                    }
            elif comp_val == "TOT" and current_subject:
                current_subject["tot_col"] = c

        if current_subject:
            subject_blocks.append(current_subject)

        if seat_col is None:
            seat_col = 2  # default
        if name_col is None:
            name_col = 3

        # Find data start
        data_start = max_row + 1

        # Match subject headers to curriculum
        for block in subject_blocks:
            header = block["header"]
            matched_sub = None
            for code, sub in sub_map.items():
                if code in header or sub.subject_name.lower() in header.lower():
                    matched_sub = sub
                    break
            if not matched_sub:
                # Try fuzzy
                for code, sub in sub_map.items():
                    if any(word in header.lower() for word in sub.subject_name.lower().split()[:2]):
                        matched_sub = sub
                        break
            block["matched_sub"] = matched_sub

        # Parse data rows
        results: List[ParsedStudentMarks] = []
        for r in range(data_start, ws.max_row + 1):
            roll_val = ws.cell(row=r, column=seat_col).value
            if roll_val is None or str(roll_val).strip() in ("", "nan", "None"):
                continue

            roll = str(roll_val).strip()
            name = str(ws.cell(row=r, column=name_col).value or "").strip()
            stu = ParsedStudentMarks(roll_number=roll, student_name=name)

            for block in subject_blocks:
                tpl = block.get("matched_sub")
                if not tpl:
                    stu.warnings.append(f"Unmatched subject: {block['header']}")
                    continue

                comp_marks = {}
                for comp_key, comp_info in block["components"].items():
                    val = ws.cell(row=r, column=comp_info["col"]).value
                    try:
                        comp_marks[comp_key] = float(val) if val is not None else 0.0
                    except (ValueError, TypeError):
                        comp_marks[comp_key] = 0.0
                        stu.errors.append(f"{tpl.subject_code}: invalid {comp_key}")

                internal, external = components_to_subject_score(comp_marks, tpl)
                total = internal + external
                gi = calculate_grade(total, tpl.internal_max + tpl.external_max)

                stu.subjects.append({
                    "subject_code": tpl.subject_code,
                    "subject_name": tpl.subject_name,
                    "credits": tpl.credits,
                    "internal_marks": internal,
                    "external_marks": external,
                    "internal_max": tpl.internal_max,
                    "external_max": tpl.external_max,
                    "total_marks": total,
                    "grade": gi["grade"],
                    "grade_points": gi["points"],
                    "is_elective": tpl.is_elective,
                    "is_practical": tpl.is_practical,
                    "components": comp_marks,
                })

            if not stu.subjects:
                stu.errors.append("No subjects found")
            results.append(stu)

        return results

    # ── XLS fallback (older Excel format) ──

    def _parse_xls_fallback(
        self, content: bytes, sub_map, elec_map,
    ) -> Tuple[List[ParsedStudentMarks], str]:
        """Read .xls as DataFrame and use generic wide-format parser."""
        try:
            df = pd.read_excel(io.BytesIO(content), header=None, engine="xlrd")
        except ImportError:
            raise ValueError("xlrd not installed — cannot read .xls files")

        # Find header row
        hdr_row = 0
        for idx, row in df.iterrows():
            vals = [str(v).strip().upper() for v in row.values if pd.notna(v)]
            if any(v in ("CA", "MSE", "ESE", "IA", "PR") for v in vals):
                hdr_row = idx
                break

        df = pd.read_excel(
            io.BytesIO(content), header=hdr_row, engine="xlrd"
        )
        return self._parse_flat_df(df, sub_map, elec_map), "xls_flat"

    # ── CSV parser ──

    def _parse_csv(
        self, content: bytes, sub_map, elec_map,
    ) -> Tuple[List[ParsedStudentMarks], str]:
        df = pd.read_csv(io.BytesIO(content))
        return self._parse_flat_df(df, sub_map, elec_map), "csv"

    # ── Flat DataFrame parser (CSV or flat XLS) ──

    def _parse_flat_df(
        self, df: pd.DataFrame, sub_map, elec_map,
    ) -> List[ParsedStudentMarks]:
        """
        Parse a flat DataFrame.
        Expects columns like:  MATH301_CA, MATH301_MSE, MATH301_ESE
        OR long format:        roll_number, subject_code, CA, MSE, ESE
        """
        cols_lower = [str(c).strip().lower() for c in df.columns]

        # Detect long format
        if "subject_code" in cols_lower and ("ca" in cols_lower or "internal_marks" in cols_lower):
            return self._parse_long_df(df, sub_map, elec_map)

        # Wide format: find component columns by pattern
        roll_col = name_col = None
        for i, cl in enumerate(cols_lower):
            if roll_col is None and ("roll" in cl or "seat" in cl):
                roll_col = i
            if name_col is None and cl in ("name", "student_name", "name of student", "student"):
                name_col = i

        if roll_col is None:
            roll_col = 0
        if name_col is None:
            name_col = 1

        # Map columns: CODE_COMPONENT pattern
        comp_cols: Dict[str, Dict[str, int]] = {}
        for i, col in enumerate(df.columns):
            cn = str(col).strip()
            for sep in ("_",):
                parts = cn.rsplit(sep, 1)
                if len(parts) == 2:
                    prefix, suffix = parts[0].strip().upper(), parts[1].strip().upper()
                    if suffix in ("CA", "MSE", "ESE", "IA", "PR", "TW", "INT", "EXT", "TOT"):
                        if prefix not in comp_cols:
                            comp_cols[prefix] = {}
                        if suffix != "TOT":
                            comp_cols[prefix][suffix] = i

        results: List[ParsedStudentMarks] = []
        df = df.dropna(how="all")

        for _, row in df.iterrows():
            rv = row.iloc[roll_col]
            roll = str(rv).strip() if pd.notna(rv) else ""
            if not roll or roll.lower() in ("nan", "none", ""):
                continue

            nm = str(row.iloc[name_col]).strip() if name_col and pd.notna(row.iloc[name_col]) else ""
            stu = ParsedStudentMarks(roll_number=roll, student_name=nm)

            for prefix, comp_indices in comp_cols.items():
                tpl = sub_map.get(prefix)
                if not tpl and prefix in elec_map:
                    einfo = elec_map[prefix]
                    tpl = einfo["template"]
                if not tpl:
                    stu.warnings.append(f"Unknown subject {prefix}")
                    continue

                comp_marks = {}
                for comp_key, col_idx in comp_indices.items():
                    val = row.iloc[col_idx]
                    comp_marks[comp_key] = float(val) if pd.notna(val) else 0.0

                # If we have INT/EXT instead of CA/MSE/ESE, convert
                if "INT" in comp_marks and "CA" not in comp_marks:
                    comp_marks["CA"] = comp_marks.pop("INT", 0)
                if "EXT" in comp_marks and "ESE" not in comp_marks:
                    ext = comp_marks.pop("EXT", 0)
                    comp_marks["MSE"] = 0
                    comp_marks["ESE"] = ext

                internal, external = components_to_subject_score(comp_marks, tpl)
                total = internal + external
                gi = calculate_grade(total, tpl.internal_max + tpl.external_max)

                stu.subjects.append({
                    "subject_code": prefix,
                    "subject_name": tpl.subject_name,
                    "credits": tpl.credits,
                    "internal_marks": internal,
                    "external_marks": external,
                    "internal_max": tpl.internal_max,
                    "external_max": tpl.external_max,
                    "total_marks": total,
                    "grade": gi["grade"],
                    "grade_points": gi["points"],
                    "is_elective": tpl.is_elective,
                    "is_practical": tpl.is_practical,
                    "components": comp_marks,
                })

            if not stu.subjects:
                stu.errors.append("No subjects found")
            results.append(stu)

        return results

    # ── Long format DataFrame parser ──

    def _parse_long_df(
        self, df: pd.DataFrame, sub_map, elec_map,
    ) -> List[ParsedStudentMarks]:
        col_map = {}
        for col in df.columns:
            n = str(col).strip().lower().replace(" ", "_")
            if "roll" in n or "seat" in n:
                col_map["roll"] = col
            elif n in ("name", "student_name", "student"):
                col_map["name"] = col
            elif n in ("subject_code", "code"):
                col_map["scode"] = col
            elif n in ("subject_name", "sub_name"):
                col_map["sname"] = col
            elif n in ("ca",):
                col_map["ca"] = col
            elif n in ("mse",):
                col_map["mse"] = col
            elif n in ("ese",):
                col_map["ese"] = col
            elif n in ("ia",):
                col_map["ia"] = col
            elif n in ("pr", "pr/or", "practical"):
                col_map["pr"] = col
            elif n in ("tw", "term_work"):
                col_map["tw"] = col
            elif n in ("internal_marks", "internal"):
                col_map["int"] = col
            elif n in ("external_marks", "external"):
                col_map["ext"] = col

        if "roll" not in col_map:
            raise ValueError("Cannot find roll/seat number column")

        bucket: Dict[str, ParsedStudentMarks] = {}

        for _, row in df.iterrows():
            rv = row.get(col_map["roll"])
            roll = str(rv).strip() if pd.notna(rv) else ""
            if not roll or roll.lower() in ("nan", "none"):
                continue

            if roll not in bucket:
                nm = str(row[col_map["name"]]).strip() if "name" in col_map and pd.notna(row.get(col_map["name"])) else ""
                bucket[roll] = ParsedStudentMarks(roll_number=roll, student_name=nm)

            stu = bucket[roll]
            sc = str(row[col_map["scode"]]).strip().upper() if "scode" in col_map and pd.notna(row.get(col_map["scode"])) else ""
            sn = str(row[col_map.get("sname", "")]).strip() if "sname" in col_map and pd.notna(row.get(col_map.get("sname", ""))) else ""

            tpl = sub_map.get(sc)
            if not tpl and sc in elec_map:
                tpl = elec_map[sc]["template"]
                sn = sn or elec_map[sc]["name"]
            if tpl:
                sn = sn or tpl.subject_name

            # Read components
            comp_marks = {}
            for key in ("ca", "mse", "ese", "ia", "pr", "tw"):
                if key in col_map:
                    val = row.get(col_map[key])
                    if pd.notna(val):
                        comp_marks[key.upper()] = float(val)

            # Fallback: internal/external columns
            if not comp_marks:
                int_val = float(row[col_map["int"]]) if "int" in col_map and pd.notna(row.get(col_map["int"])) else 0
                ext_val = float(row[col_map["ext"]]) if "ext" in col_map and pd.notna(row.get(col_map["ext"])) else 0
                internal, external = int_val, ext_val
            elif tpl:
                internal, external = components_to_subject_score(comp_marks, tpl)
            else:
                internal = comp_marks.get("CA", comp_marks.get("IA", comp_marks.get("TW", 0)))
                external = comp_marks.get("MSE", 0) + comp_marks.get("ESE", comp_marks.get("PR", 0))

            credits = tpl.credits if tpl else 3
            int_max = tpl.internal_max if tpl else 20
            ext_max = tpl.external_max if tpl else 80
            total = internal + external
            gi = calculate_grade(total, int_max + ext_max)

            stu.subjects.append({
                "subject_code": sc,
                "subject_name": sn,
                "credits": credits,
                "internal_marks": internal,
                "external_marks": external,
                "internal_max": int_max,
                "external_max": ext_max,
                "total_marks": total,
                "grade": gi["grade"],
                "grade_points": gi["points"],
                "is_elective": tpl.is_elective if tpl else False,
                "is_practical": tpl.is_practical if tpl else False,
                "components": comp_marks,
            })

        return list(bucket.values())

    # ─────────────────────────────────────────────────
    # 5.3  CSV CONVERSION
    # ─────────────────────────────────────────────────

    def to_csv(self, parsed: List[ParsedStudentMarks]) -> str:
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow([
            "roll_number", "student_name", "subject_code", "subject_name",
            "credits", "internal_marks", "external_marks", "total_marks",
            "grade", "grade_points", "is_elective", "is_practical",
            "components_json",
        ])
        for stu in parsed:
            for s in stu.subjects:
                writer.writerow([
                    stu.roll_number, stu.student_name,
                    s["subject_code"], s["subject_name"], s["credits"],
                    s["internal_marks"], s["external_marks"], s["total_marks"],
                    s["grade"], s["grade_points"], s["is_elective"], s["is_practical"],
                    json.dumps(s.get("components", {})),
                ])
        return buf.getvalue()

    # ─────────────────────────────────────────────────
    # 5.4  PREVIEW (unchanged)
    # ─────────────────────────────────────────────────

    async def preview(
        self,
        parsed: List[ParsedStudentMarks],
        semester: int, branch: str, academic_year: str,
    ) -> UploadResult:
        result = UploadResult(
            upload_id=str(uuid.uuid4()),
            total_rows=len(parsed),
            semester=semester, branch=branch, academic_year=academic_year,
        )
        result.csv_data = self.to_csv(parsed)

        for stu in parsed:
            profile = await self._find_profile(stu.roll_number)
            if profile is None:
                result.unmatched_students += 1
                result.unmatched_roll_numbers.append(stu.roll_number)
                continue

            result.matched_students += 1
            tgp = tc = ce = 0.0
            for s in stu.subjects:
                c = s["credits"]
                tc += c
                if s["grade"] != "F":
                    tgp += s["grade_points"] * c
                    ce += c
            # Use SGPI from sheet if available
            calculated_sgpa = round(tgp / tc, 2) if tc > 0 else 0.0
            sgpa = stu.sgpa_from_sheet if stu.sgpa_from_sheet is not None and stu.sgpa_from_sheet > 0 else calculated_sgpa

            existing = any(sr.semester_number == semester for sr in profile.semester_records)

            result.matched_details.append({
                "roll_number": stu.roll_number,
                "student_name": stu.student_name or profile.name,
                "profile_name": profile.name,
                "user_id": profile.user_id,
                "branch": profile.branch,
                "current_cgpa": profile.cgpa,
                "preview_sgpa": sgpa,
                "credits_earned": int(ce),
                "total_credits": int(tc),
                "subjects_count": len(stu.subjects),
                "subjects": stu.subjects,
                "errors": stu.errors,
                "warnings": stu.warnings,
                "has_errors": bool(stu.errors),
                "has_existing_semester": existing,
            })

        return result

    # ─────────────────────────────────────────────────
    # 5.5  SAVE (REPLACED with updated version from patch)
    # ─────────────────────────────────────────────────

    async def save(
        self,
        parsed: List[ParsedStudentMarks],
        semester: int, academic_year: str, branch: str,
        overwrite: bool = True,
        admin_email: str = None,
    ) -> UploadResult:
        result = UploadResult(
            upload_id=str(uuid.uuid4()),
            total_rows=len(parsed),
            semester=semester, branch=branch, academic_year=academic_year,
        )
        result.csv_data = self.to_csv(parsed)

        for stu in parsed:
            # Always save/update pending marks as backup (even for matched students)
            try:
                await self._save_pending_marks(
                    stu, semester, academic_year, branch, admin_email
                )
            except Exception as e:
                logger.warning(f"Backup pending save failed for {stu.roll_number}: {e}")

            profile = await self._find_profile(stu.roll_number)

            if profile is None:
                result.unmatched_students += 1
                result.unmatched_roll_numbers.append(stu.roll_number)
                result.created_students += 1
                logger.info(f"Saved pending marks for unregistered student: {stu.roll_number}")
                continue

            # ── Process matched students ──
            result.matched_students += 1

            if stu.errors:
                result.failed_updates += 1
                result.errors.append({
                    "roll_number": stu.roll_number, "errors": stu.errors,
                })
                continue

            try:
                processed = []
                tgp = tc = ce = 0.0

                for sd in stu.subjects:
                    ss = SubjectScore(
                        subject_code=sd["subject_code"],
                        subject_name=sd["subject_name"],
                        credits=sd["credits"],
                        internal_marks=sd["internal_marks"],
                        external_marks=sd["external_marks"],
                        total_marks=sd["total_marks"],
                        grade=sd["grade"],
                        grade_points=sd["grade_points"],
                        is_elective=sd["is_elective"],
                        is_practical=sd["is_practical"],
                    )
                    processed.append(ss)
                    c = sd["credits"]
                    tc += c
                    if sd["grade"] != "F":
                        tgp += sd["grade_points"] * c
                        ce += c

                # Use SGPI from sheet if available (more accurate than recalculating
                # since the sheet uses the original MU grading scale)
                calculated_sgpa = round(tgp / tc, 2) if tc > 0 else 0.0
                sgpa = stu.sgpa_from_sheet if stu.sgpa_from_sheet is not None and stu.sgpa_from_sheet > 0 else calculated_sgpa

                sem_rec = SemesterRecord(
                    semester_number=semester,
                    academic_year=academic_year,
                    subjects=processed,
                    sgpa=sgpa,
                    total_credits=int(tc),
                    credits_earned=int(ce),
                    is_complete=True,
                    created_at=datetime.now(),
                )

                idx = next(
                    (i for i, sr in enumerate(profile.semester_records)
                     if sr.semester_number == semester),
                    None,
                )
                if idx is not None:
                    if not overwrite:
                        result.skipped_students += 1
                        result.warnings.append(
                            f"{stu.roll_number}: sem {semester} exists — skipped"
                        )
                        continue
                    profile.semester_records[idx] = sem_rec
                else:
                    profile.semester_records.append(sem_rec)
                    profile.semester_records.sort(key=lambda x: x.semester_number)

                # Use CGPI from sheet if available; otherwise recalculate from semester SGPAs
                if stu.cgpa_from_sheet is not None and stu.cgpa_from_sheet > 0:
                    profile.cgpa = round(stu.cgpa_from_sheet, 2)
                else:
                    agp = ac = 0.0
                    for sr in profile.semester_records:
                        if sr.is_complete and sr.total_credits > 0:
                            agp += sr.sgpa * sr.total_credits
                            ac += sr.total_credits
                    profile.cgpa = round(agp / ac, 2) if ac > 0 else 0.0

                ace = sum(sr.credits_earned for sr in profile.semester_records if sr.is_complete)
                profile.total_credits_earned = int(ace)
                profile.last_updated = datetime.now()
                profile.marks_synced_at = datetime.now()

                # ✅ FIX: Update current_semester to highest completed semester
                # This ensures student_analysis filters work correctly
                completed_sems = [
                    sr.semester_number
                    for sr in profile.semester_records
                    if sr.is_complete
                ]
                if completed_sems:
                    profile.current_semester = max(completed_sems)

                # Reset pending_marks_checked so the student's next login picks up changes
                profile.pending_marks_checked = False

                await profile.replace()

                result.updated_students += 1
                result.matched_details.append({
                    "roll_number": stu.roll_number,
                    "student_name": profile.name,
                    "profile_name": profile.name,
                    "user_id": profile.user_id,
                    "branch": profile.branch,
                    "current_cgpa": profile.cgpa,
                    "preview_sgpa": sgpa,
                    "credits_earned": int(ce),
                    "total_credits": int(tc),
                    "subjects_count": len(processed),
                    "subjects": stu.subjects,
                    "errors": [],
                    "warnings": stu.warnings,
                    "has_errors": False,
                    "updated_cgpa": profile.cgpa,
                    "status": "updated",
                })

                logger.info(
                    f"✅ Updated marks for {stu.roll_number}: "
                    f"Sem {semester} SGPA={sgpa} (sheet={stu.sgpa_from_sheet}), "
                    f"CGPA={profile.cgpa} (sheet={stu.cgpa_from_sheet})"
                )

            except Exception as e:
                result.failed_updates += 1
                result.errors.append({"roll_number": stu.roll_number, "error": str(e)})
                logger.error(f"Save failed for {stu.roll_number}: {e}", exc_info=True)

        return result

    # ─────────────────────────────────────────────────
    # 5.6  HELPERS (UPDATED _save_pending_marks and _find_profile)
    # ─────────────────────────────────────────────────

    async def _save_pending_marks(
        self,
        stu: ParsedStudentMarks,
        semester: int,
        academic_year: str,
        branch: str,
        admin_email: str = None,
        seat_number: Optional[str] = None,
    ) -> None:
        """
        Save marks to pending_student_marks collection.
        FIXED: Saves old linked_user_id before resetting so profile flag gets cleared.
        """
        processed_subjects = []
        tgp = tc = ce = 0.0

        for sd in stu.subjects:
            ss = SubjectScore(
                subject_code=sd["subject_code"],
                subject_name=sd["subject_name"],
                credits=sd["credits"],
                internal_marks=sd["internal_marks"],
                external_marks=sd["external_marks"],
                total_marks=sd["total_marks"],
                grade=sd["grade"],
                grade_points=sd["grade_points"],
                is_elective=sd["is_elective"],
                is_practical=sd["is_practical"],
            )
            processed_subjects.append(ss)
            c = sd["credits"]
            tc += c
            if sd["grade"] != "F":
                tgp += sd["grade_points"] * c
                ce += c

        # Use SGPI from sheet if available
        calculated_sgpa = round(tgp / tc, 2) if tc > 0 else 0.0
        sgpa = stu.sgpa_from_sheet if stu.sgpa_from_sheet is not None and stu.sgpa_from_sheet > 0 else calculated_sgpa

        # Extract admission year from roll number
        admission_year_from_roll = None
        if stu.roll_number and len(stu.roll_number) >= 4:
            try:
                admission_year_from_roll = int(stu.roll_number[:4])
            except ValueError:
                pass

        identifier = stu.roll_number
        detected_seat_number = seat_number
        if identifier and len(identifier) >= 5 and len(identifier) <= 6 and identifier.isdigit():
            detected_seat_number = identifier[-5:]

        # ── Find existing pending marks ──
        or_conditions = [
            {"roll_number": stu.roll_number, "semester_number": semester}
        ]
        if detected_seat_number:
            or_conditions.append(
                {"seat_number": detected_seat_number, "semester_number": semester}
            )

        existing = await PendingStudentMarks.find_one({"$or": or_conditions})

        if existing:
            # Save old linked info BEFORE resetting
            old_linked_user_id = existing.linked_user_id
            old_was_linked = existing.linked_to_profile

            existing.subjects = processed_subjects
            existing.sgpa = sgpa
            existing.total_credits = int(tc)
            existing.credits_earned = int(ce)
            existing.academic_year = academic_year
            existing.upload_timestamp = datetime.now()
            existing.uploaded_by = admin_email or "admin"
            existing.linked_to_profile = False
            existing.linked_user_id = None
            if detected_seat_number:
                existing.seat_number = detected_seat_number
            await existing.replace()

            # Reset the OLD linked profile's flag using the saved value
            if old_was_linked and old_linked_user_id:
                try:
                    linked_profile = await StudentProfile.find_one(
                        StudentProfile.user_id == old_linked_user_id
                    )
                    if linked_profile:
                        linked_profile.pending_marks_checked = False
                        await linked_profile.replace()
                        logger.info(
                            f"Reset pending_marks_checked for previously linked user: "
                            f"{old_linked_user_id}"
                        )
                except Exception as e:
                    logger.warning(f"Could not reset linked profile flag: {e}")

            # Also try to find and reset by roll number
            try:
                profile_by_roll = await StudentProfile.find_one(
                    StudentProfile.roll_number == stu.roll_number
                )
                if profile_by_roll:
                    profile_by_roll.pending_marks_checked = False
                    await profile_by_roll.replace()
            except Exception as e:
                logger.warning(f"Could not reset profile by roll: {e}")

            logger.info(f"Updated pending marks for {stu.roll_number} sem {semester}")

        else:
            # Create new pending marks entry
            pending = PendingStudentMarks(
                roll_number=identifier,
                seat_number=detected_seat_number,
                student_name=stu.student_name,
                branch=branch,
                admission_year=admission_year_from_roll or 2022,
                semester_number=semester,
                academic_year=academic_year,
                subjects=processed_subjects,
                sgpa=sgpa,
                total_credits=int(tc),
                credits_earned=int(ce),
                uploaded_by=admin_email or "admin",
            )
            await pending.insert()
            logger.info(f"Created pending marks for {stu.roll_number} sem {semester}")

    @staticmethod
    async def _find_profile(identifier: str) -> Optional[StudentProfile]:
        """
        Find profile by roll number OR seat number.
        FIXED: Prefer profiles with real Firebase UIDs over placeholders.
        """
        if not identifier or not identifier.strip():
            return None

        identifier = identifier.strip()

        # ── 1. Exact roll number match ──
        profiles = await StudentProfile.find(
            StudentProfile.roll_number == identifier
        ).to_list()

        if profiles:
            # Prefer profile with real user_id (not placeholder from add_students_batch)
            real = [p for p in profiles if not p.user_id.startswith("pending_")]
            if real:
                return real[0]
            return profiles[0]

        # ── 2. Case-insensitive roll number match ──
        profiles = await StudentProfile.find({
            "roll_number": {"$regex": f"^{re.escape(identifier)}$", "$options": "i"}
        }).to_list()

        if profiles:
            real = [p for p in profiles if not p.user_id.startswith("pending_")]
            if real:
                return real[0]
            return profiles[0]

        # ── 3. Seat number match (current) ──
        p = await StudentProfile.find_one(
            StudentProfile.current_seat_number == identifier
        )
        if p:
            return p

        # ── 4. Seat number match (history) ──
        p = await StudentProfile.find_one({
            "seat_number_history.seat_number": identifier
        })
        if p:
            return p

        # ── 5. Fuzzy roll number match (single result only) ──
        hits = await StudentProfile.find({
            "roll_number": {"$regex": re.escape(identifier), "$options": "i"}
        }).to_list()
        return hits[0] if len(hits) == 1 else None

    # ─────────────────────────────────────────────────
    # 5.7  TEMPLATE WITH PRE-FILLED STUDENTS (existing)
    # ─────────────────────────────────────────────────

    async def generate_template_with_students(
        self,
        semester: int,
        branch: str,
        academic_year: str,
        admission_year: int,
        elective_choices: Optional[Dict[str, str]] = None,
        prefill_branch: Optional[str] = None,
        prefill_semester: Optional[int] = None,
    ) -> io.BytesIO:
        """
        Generate template AND pre-fill with existing student roll numbers.

        Queries student_profiles for matching branch and optionally filters
        by current_semester or admission_year.
        """
        # Step 1: Generate the normal template
        buf = self.generate_template(
            semester=semester,
            branch=branch,
            academic_year=academic_year,
            admission_year=admission_year,
            elective_choices=elective_choices,
        )

        # Step 2: Fetch matching students from DB
        query_filter: Dict[str, Any] = {}
        target_branch = prefill_branch or branch
        if target_branch:
            query_filter["branch"] = {"$regex": f"^{target_branch}$", "$options": "i"}
        if admission_year:
            query_filter["admission_year"] = admission_year

        students = await StudentProfile.find(query_filter).sort("roll_number").to_list()

        if not students:
            logger.warning(f"No students found for branch={target_branch}, admission={admission_year}")
            return buf

        # Step 3: Open the workbook and fill in student data
        from openpyxl import load_workbook as _lw

        wb = _lw(buf)
        ws = wb["Marks Data"]

        # Find the data start row
        data_start = 7
        for r in range(4, 20):
            v = ws.cell(row=r, column=1).value
            if v is not None and str(v).strip().isdigit():
                data_start = r
                break

        # Fill student rows
        for i, student in enumerate(students):
            row = data_start + i
            ws.cell(row=row, column=1, value=i + 1)  # Sr. No
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=2, value=student.roll_number)  # Seat No
            ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=2).font = Font(bold=True)
            ws.cell(row=row, column=3, value=student.name)  # Name
            ws.cell(row=row, column=3).alignment = Alignment(horizontal="left")

            # Apply borders to all columns in this row
            for c in range(1, ws.max_column + 1):
                ws.cell(row=row, column=c).border = _BORDER

        # Store student count metadata
        meta_row = 101
        ws.cell(row=meta_row, column=1, value="__STUDENTS_META__")
        ws.cell(row=meta_row, column=2, value=json.dumps({
            "total_students": len(students),
            "branch": target_branch,
            "admission_year": admission_year,
        }))
        ws.row_dimensions[meta_row].hidden = True

        # Save back
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out

    @staticmethod
    async def get_students_for_template(
        branch: str,
        admission_year: Optional[int] = None,
    ) -> List[Dict[str, Any]]:
        """
        Fetch student roll numbers and names for pre-filling templates.
        """
        query_filter: Dict[str, Any] = {}
        if branch:
            query_filter["branch"] = {"$regex": f"^{branch}$", "$options": "i"}
        if admission_year:
            query_filter["admission_year"] = admission_year

        students = await StudentProfile.find(query_filter).sort("roll_number").to_list()

        return [
            {
                "roll_number": s.roll_number,
                "name": s.name,
                "user_id": s.user_id,
                "branch": s.branch,
                "admission_year": s.admission_year,
                "current_semester": s.current_semester,
                "cgpa": s.cgpa,
            }
            for s in students
        ]

    # ─────────────────────────────────────────────────
    # 5.8  ADDITIONAL METHODS FROM PATCH
    # ─────────────────────────────────────────────────

    async def generate_template_with_marks(
        self,
        semester: int,
        branch: str,
        academic_year: str,
        admission_year: int,
        elective_choices: Optional[Dict[str, str]] = None,
    ) -> io.BytesIO:
        """
        Generate an XLSX template with existing marks data pre-filled.
        Fetches students and their semester records from the database.
        """
        subjects = get_semester_subjects(semester, admission_year)
        if not subjects:
            raise ValueError(
                f"No subjects for semester {semester}, admission {admission_year}"
            )

        # Resolve elective choices
        resolved = self._resolve_electives(subjects, elective_choices)

        # Fetch students with their marks for this semester
        query_filter: Dict[str, Any] = {}
        if branch:
            query_filter["branch"] = {"$regex": f"^{branch}$", "$options": "i"}
        if admission_year:
            query_filter["admission_year"] = admission_year

        students = await StudentProfile.find(query_filter).sort("roll_number").to_list()

        # Build student marks data
        students_with_marks = []
        for student in students:
            # Find semester record for this semester
            sem_record = None
            for sr in student.semester_records:
                if sr.semester_number == semester:
                    sem_record = sr
                    break

            student_data = {
                "roll_number": student.roll_number,
                "name": student.name,
                "seat_number": student.current_seat_number or "",
                "subjects": {},
                "sgpa": sem_record.sgpa if sem_record else None,
                "total_marks": 0,
                "has_marks": sem_record is not None,
            }

            if sem_record:
                # Map subject marks
                for subj in sem_record.subjects:
                    student_data["subjects"][subj.subject_code] = {
                        "internal_marks": subj.internal_marks,
                        "external_marks": subj.external_marks,
                        "total_marks": subj.total_marks,
                        "grade": subj.grade,
                        "grade_points": subj.grade_points,
                    }
                    student_data["total_marks"] += subj.total_marks

            students_with_marks.append(student_data)

        # Generate workbook
        wb = Workbook()
        self._build_marks_sheet_with_data(
            wb, resolved, semester, branch, academic_year, admission_year,
            students_with_marks
        )
        self._build_info_sheet(wb, resolved)

        elec = [s for s in subjects if s.is_elective and s.elective_group]
        if elec:
            self._build_elective_sheet(wb, elec)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def _build_marks_sheet_with_data(
        self, wb, subjects, semester, branch, ay, adm_yr, students_with_marks
    ):
        """Build marks sheet with existing marks data pre-filled"""
        ws = wb.active
        ws.title = "Marks Data"
        S = STYLE

        # ─── Row 1: metadata ───
        meta = [
            ("semester", semester), ("branch", branch),
            ("academic_year", ay), ("admission_year", adm_yr),
        ]
        c = 1
        for key, val in meta:
            k_cell = ws.cell(row=1, column=c, value=key)
            k_cell.font = S["meta_font"]; k_cell.fill = S["meta_fill"]
            v_cell = ws.cell(row=1, column=c + 1, value=val)
            v_cell.font = S["meta_font"]; v_cell.fill = S["meta_fill"]
            c += 2

        # ─── Row 2: instructions ───
        ws.cell(row=2, column=1, value=(
            "This template contains existing marks data. "
            "Modify as needed and re-upload to update records."
        )).font = S["inst_font"]
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=14)

        HDR = 4      # subject name row
        COMP = 5     # component row
        MAX_R = 6    # max-marks row
        DATA = 7     # first data row

        # ─── Fixed columns ───
        fixed_headers = [
            ("Sr.\nNo", 6),
            ("Seat No\n(Roll Number)", 20),
            ("Name of\nStudent", 30),
        ]
        for i, (label, width) in enumerate(fixed_headers, 1):
            cell = ws.cell(row=HDR, column=i, value=label)
            cell.font = S["hdr_font"]; cell.fill = S["hdr_fill"]
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = S["border"]
            ws.merge_cells(start_row=HDR, start_column=i, end_row=COMP, end_column=i)
            for r in (HDR, COMP):
                ws.cell(row=r, column=i).border = S["border"]
            ws.column_dimensions[get_column_letter(i)].width = width

        # ─── Subject columns ───
        col = len(fixed_headers) + 1
        column_map: List[Dict[str, Any]] = []
        subject_col_mapping = {}  # Maps subject_code -> {component -> col}

        for sub in subjects:
            comps = get_subject_components(sub)
            n_comp = len(comps)
            n_cols = n_comp + 1  # +1 for TOT

            has_elective_code = sub.is_elective and sub.elective_group
            if has_elective_code:
                n_cols += 1

            start_col = col
            subject_col_mapping[sub.subject_code] = {"start": start_col, "components": {}}

            # ── Subject header ──
            type_label = get_subject_type_label(sub)
            if has_elective_code:
                header_text = (
                    f"{sub.elective_group}\n({sub.subject_code}) "
                    f"[{sub.credits} cr, {type_label}]"
                )
            else:
                header_text = (
                    f"{sub.subject_name}\n({sub.subject_code}) "
                    f"[{sub.credits} cr, {type_label}]"
                )

            cell = ws.cell(row=HDR, column=start_col, value=header_text)
            fill = CATEGORY_FILLS.get(sub.course_type, S["hdr_fill"])
            cell.font = S["hdr_font"]; cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = S["border"]

            if n_cols > 1:
                ws.merge_cells(
                    start_row=HDR, start_column=start_col,
                    end_row=HDR, end_column=start_col + n_cols - 1,
                )

            # ── Component headers ──
            ci = start_col

            if has_elective_code:
                cc = ws.cell(row=COMP, column=ci, value="ELEC\nCODE")
                cc.font = S["comp_font"]; cc.fill = S["elec_fill"]
                cc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cc.border = S["border"]
                mc = ws.cell(row=MAX_R, column=ci, value="Code")
                mc.font = S["max_font"]; mc.fill = S["max_fill"]
                mc.alignment = Alignment(horizontal="center")
                mc.border = S["border"]
                ws.column_dimensions[get_column_letter(ci)].width = 14
                column_map.append({
                    "col": ci, "subject_code": sub.subject_code,
                    "component": "ELEC_CODE", "max": 0,
                    "elective_group": sub.elective_group,
                })
                subject_col_mapping[sub.subject_code]["elec_code_col"] = ci
                ci += 1

            for comp in comps:
                cc = ws.cell(row=COMP, column=ci, value=comp["label"])
                cc.font = S["comp_font"]; cc.fill = S["comp_fill"]
                cc.alignment = Alignment(horizontal="center", vertical="center")
                cc.border = S["border"]

                mc = ws.cell(row=MAX_R, column=ci, value=comp["max"])
                mc.font = S["max_font"]; mc.fill = S["max_fill"]
                mc.alignment = Alignment(horizontal="center")
                mc.border = S["border"]

                ws.column_dimensions[get_column_letter(ci)].width = 8
                column_map.append({
                    "col": ci, "subject_code": sub.subject_code,
                    "component": comp["key"], "max": comp["max"],
                    "elective_group": sub.elective_group or "",
                })
                subject_col_mapping[sub.subject_code]["components"][comp["key"]] = ci
                ci += 1

            # TOT column
            tc = ws.cell(row=COMP, column=ci, value="TOT")
            tc.font = Font(bold=True, size=10); tc.fill = S["comp_fill"]
            tc.alignment = Alignment(horizontal="center")
            tc.border = S["border"]
            total_max = sum(comp["max"] for comp in comps)
            tm = ws.cell(row=MAX_R, column=ci, value=total_max)
            tm.font = S["max_font"]; tm.fill = S["max_fill"]
            tm.alignment = Alignment(horizontal="center")
            tm.border = S["border"]
            ws.column_dimensions[get_column_letter(ci)].width = 8
            column_map.append({
                "col": ci, "subject_code": sub.subject_code,
                "component": "TOT", "max": total_max,
                "elective_group": sub.elective_group or "",
            })
            subject_col_mapping[sub.subject_code]["tot_col"] = ci
            ci += 1

            col = ci

        # ── Summary columns ──
        summary_start_col = col
        summary_headers = ["Total\nMarks", "SGPI", "Result"]
        for label in summary_headers:
            cell = ws.cell(row=HDR, column=col, value=label)
            cell.font = S["hdr_font"]; cell.fill = S["summary_fill"]
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = S["border"]
            ws.merge_cells(start_row=HDR, start_column=col, end_row=COMP, end_column=col)
            ws.column_dimensions[get_column_letter(col)].width = 10
            col += 1

        # ── Max row styling for fixed cols ──
        for i in range(1, len(fixed_headers) + 1):
            c = ws.cell(row=MAX_R, column=i, value="")
            c.fill = S["max_fill"]; c.border = S["border"]

        # ═══════════════════════════════════════════════════
        # FILL STUDENT DATA WITH MARKS
        # ═══════════════════════════════════════════════════
        total_cols = col - 1

        for idx, student in enumerate(students_with_marks):
            row = DATA + idx

            # Sr. No
            ws.cell(row=row, column=1, value=idx + 1).alignment = Alignment(horizontal="center")

            # Roll Number / Seat No
            roll_cell = ws.cell(row=row, column=2, value=student["roll_number"])
            roll_cell.alignment = Alignment(horizontal="center")
            roll_cell.font = Font(bold=True)

            # Name
            ws.cell(row=row, column=3, value=student["name"]).alignment = Alignment(horizontal="left")

            # Fill subject marks
            for sub in subjects:
                sub_code = sub.subject_code
                if sub_code not in subject_col_mapping:
                    continue

                col_info = subject_col_mapping[sub_code]

                # Check if student has marks for this subject
                if sub_code in student["subjects"]:
                    marks_data = student["subjects"][sub_code]
                    internal = marks_data["internal_marks"]
                    external = marks_data["external_marks"]
                    total = marks_data["total_marks"]

                    # Get subject components to properly split marks
                    comps = get_subject_components(sub)

                    # Fill component columns based on subject type
                    if sub.course_type in ("MNP", "MJP", "INT"):
                        # Project/Internship: TW only
                        if "TW" in col_info["components"]:
                            ws.cell(row=row, column=col_info["components"]["TW"], value=internal)

                    elif sub.course_type == "SBL":
                        # Skill Lab: TW + PR
                        if "TW" in col_info["components"]:
                            ws.cell(row=row, column=col_info["components"]["TW"], value=internal)
                        if "PR" in col_info["components"]:
                            ws.cell(row=row, column=col_info["components"]["PR"], value=external)

                    elif sub.course_type == "LBC" or sub.is_practical:
                        # Lab: IA + PR
                        if "IA" in col_info["components"]:
                            ws.cell(row=row, column=col_info["components"]["IA"], value=internal)
                        if "PR" in col_info["components"]:
                            ws.cell(row=row, column=col_info["components"]["PR"], value=external)

                    else:
                        # Theory: CA + MSE + ESE
                        # internal = CA, external = MSE + ESE
                        if "CA" in col_info["components"]:
                            ws.cell(row=row, column=col_info["components"]["CA"], value=internal)

                        # Split external marks into MSE and ESE
                        # Default ratio: MSE=30, ESE=50 (total external=80)
                        if "MSE" in col_info["components"] and "ESE" in col_info["components"]:
                            # Try to estimate MSE/ESE split
                            # If external_max is 80, assume MSE=30, ESE=50
                            ext_max = sub.external_max
                            if ext_max == 80:
                                mse_max, ese_max = 30, 50
                            else:
                                mse_max = round(ext_max * 0.375)
                                ese_max = ext_max - mse_max

                            # Proportionally split actual marks
                            if ext_max > 0:
                                mse_marks = round(external * (mse_max / ext_max), 1)
                                ese_marks = round(external - mse_marks, 1)
                            else:
                                mse_marks, ese_marks = 0, 0

                            ws.cell(row=row, column=col_info["components"]["MSE"], value=mse_marks)
                            ws.cell(row=row, column=col_info["components"]["ESE"], value=ese_marks)

                    # Fill TOT column
                    if "tot_col" in col_info:
                        ws.cell(row=row, column=col_info["tot_col"], value=total)

            # Fill summary columns
            if student["has_marks"]:
                # Total Marks
                ws.cell(row=row, column=summary_start_col, value=student["total_marks"])
                # SGPI
                sgpi_cell = ws.cell(row=row, column=summary_start_col + 1, value=student["sgpa"])
                sgpi_cell.font = Font(bold=True)
                # Result
                result = "PASS" if student["sgpa"] and student["sgpa"] >= 4.0 else "FAIL" if student["sgpa"] else ""
                ws.cell(row=row, column=summary_start_col + 2, value=result)

            # Apply borders to all columns
            for c2 in range(1, total_cols + 1):
                ws.cell(row=row, column=c2).border = S["border"]
                if c2 > 3:  # Mark columns
                    ws.cell(row=row, column=c2).alignment = Alignment(horizontal="center")

        # ── Store column map ──
        ws.cell(row=100, column=1, value="__COLUMN_MAP__")
        ws.cell(row=100, column=2, value=json.dumps(column_map))
        ws.row_dimensions[100].hidden = True

        # ── Store students metadata ──
        ws.cell(row=101, column=1, value="__STUDENTS_META__")
        ws.cell(row=101, column=2, value=json.dumps({
            "total_students": len(students_with_marks),
            "students_with_marks": len([s for s in students_with_marks if s["has_marks"]]),
            "branch": branch,
            "admission_year": adm_yr,
            "semester": semester,
        }))
        ws.row_dimensions[101].hidden = True

        # ── Freeze panes ──
        ws.freeze_panes = f"D{DATA}"

        # ── Color rows based on marks status ──
        for idx, student in enumerate(students_with_marks):
            row = DATA + idx
            if student["has_marks"]:
                # Light green for students with existing marks
                for c2 in range(1, 4):
                    ws.cell(row=row, column=c2).fill = PatternFill("solid", fgColor="E2EFDA")

    async def auto_register_student_from_marks(
        self,
        roll_number: str,
        name: str,
        seat_number: str,
        branch: str,
        semester: int,
        admission_year: int,
        academic_year: str
    ) -> Dict[str, Any]:
        """
        Auto-register a student when marks are uploaded.
        Called internally when a new roll number is found in marks Excel.
        """
        from app.database.connection import get_mongo_database

        db = get_mongo_database()
        if not db:
            return {"success": False, "error": "Database unavailable"}

        try:
            # Check if already exists
            existing = await db.student_profiles.find_one({"roll_number": roll_number})
            if existing:
                return {"success": True, "student_id": str(existing["_id"]), "already_exists": True}

            # Generate password
            default_password = generate_student_password(roll_number, admission_year)
            password_hash = hash_password(default_password)

            # Extract email from college pattern
            email = f"{roll_number}@college.edu"

            # Create profile
            profile_doc = {
                "name": name,
                "roll_number": roll_number,
                "seat_number": seat_number if seat_number else None,
                "email": email,
                "branch": branch.upper(),
                "admission_year": admission_year,
                "current_semester": semester,
                "current_academic_year": academic_year,
                "password_hash": password_hash,
                "password_changed": False,
                "cgpa": 0.0,
                "total_credits_earned": 0,
                "total_credits_required": 160,
                "semester_records": [],
                "created_at": datetime.utcnow(),
                "last_updated": datetime.utcnow(),
                "created_by": "auto_marks_upload",
                "auto_registered": True,  # Flag for tracking
            }

            result = await db.student_profiles.insert_one(profile_doc)

            logger.info(f"✅ Auto-registered student: {roll_number} ({name})")

            return {
                "success": True,
                "student_id": str(result.inserted_id),
                "roll_number": roll_number,
                "name": name,
                "default_password": default_password,
                "email": email,
                "auto_registered": True
            }

        except Exception as e:
            logger.error(f"Auto-registration failed for {roll_number}: {e}")
            return {"success": False, "error": str(e)}


# ══════════════════════════════════════════════════════════
bulk_marks_service = BulkMarksService()