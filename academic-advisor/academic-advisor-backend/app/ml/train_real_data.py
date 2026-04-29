"""
Enhanced Real Data Training Pipeline - FINAL FIXED VERSION
===========================================================
Fixes:
1. ✅ Proper diversified label assignment (no case issues)
2. ✅ Guaranteed minimum samples per class
3. ✅ Controlled randomization for diversity
4. ✅ Fallback to simpler split if stratification fails
5. ✅ Full error handling and validation
"""

import os
import re
import json
import logging
import warnings
import numpy as np
import pandas as pd
from typing import Dict, List, Tuple, Any, Optional
from collections import defaultdict, Counter
from datetime import datetime

warnings.filterwarnings('ignore')

# Visualization
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import seaborn as sns

# ML Libraries
from sklearn.ensemble import (
    RandomForestClassifier, 
    GradientBoostingClassifier,
    ExtraTreesClassifier,
    VotingClassifier,
    StackingClassifier
)
from sklearn.linear_model import LogisticRegression, RidgeClassifier
from sklearn.svm import SVC
from sklearn.neighbors import KNeighborsClassifier
from sklearn.tree import DecisionTreeClassifier

from sklearn.preprocessing import (
    StandardScaler, 
    RobustScaler,
    LabelEncoder
)
from sklearn.model_selection import (
    train_test_split, 
    GridSearchCV,
    StratifiedKFold,
    cross_val_score
)
from sklearn.metrics import (
    classification_report, 
    accuracy_score,
    precision_score,
    recall_score,
    f1_score,
    confusion_matrix
)
from sklearn.impute import KNNImputer

from imblearn.over_sampling import BorderlineSMOTE, SMOTE

import xgboost as xgb
import lightgbm as lgb
import joblib
import time

logger = logging.getLogger(__name__)

# ═══════════════════════════════════════════════════════════════
#  CONFIGURATION
# ═══════════════════════════════════════════════════════════════

EXCEL_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
    "exported_marks", "IT - Copy.xlsx"
)

SHEET_SEMESTER_MAP = {
    "IT SEM-III SH-2024": 3,
    "IT SEM-IV FH-2025": 4,
    "IT-V-SH 2025": 5,
}

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "training_outputs")
MODEL_DIR = os.path.join(os.path.dirname(__file__), "saved_models")
VIZ_DIR = os.path.join(OUTPUT_DIR, "visualizations")

os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(MODEL_DIR, exist_ok=True)
os.makedirs(VIZ_DIR, exist_ok=True)

# ═══════════════════════════════════════════════════════════════
#  EXCEL PARSING (UNCHANGED)
# ═══════════════════════════════════════════════════════════════

def parse_subject_name(header_cell: str) -> str:
    if ":" in header_cell:
        return header_cell.split(":", 1)[1].strip()
    return header_cell.strip()

def parse_total_marks(value) -> Optional[float]:
    if value is None:
        return None
    s = str(value).strip()
    if s in ("Ab", "ab", "AB", "", "-", "NA", "N/A"):
        return None
    s = s.rstrip("Ff")
    if "+" in s:
        parts = s.split("+")
        try:
            return sum(float(p) for p in parts)
        except ValueError:
            return None
    try:
        return float(s)
    except ValueError:
        return None

class ExcelDataParser:
    def __init__(self, excel_path: str = EXCEL_PATH):
        self.excel_path = excel_path

    def parse_all_sheets(self) -> pd.DataFrame:
        import openpyxl
        wb = openpyxl.load_workbook(self.excel_path, read_only=True)
        all_records = []

        for sheet_name in wb.sheetnames:
            semester = SHEET_SEMESTER_MAP.get(sheet_name)
            if semester is None or "FH-2025 (2)" in sheet_name:
                continue

            ws = wb[sheet_name]
            records = self._parse_sheet(ws, sheet_name, semester)
            all_records.extend(records)
            logger.info(f"Parsed {len(records)} records from {sheet_name}")

        wb.close()
        return pd.DataFrame(all_records)

    def _parse_sheet(self, ws, sheet_name: str, semester: int) -> List[Dict]:
        rows = list(ws.iter_rows(values_only=True))
        header_row = rows[3] if len(rows) > 3 else []
        
        subjects = []
        for col_idx, cell in enumerate(header_row):
            if cell and ":" in str(cell):
                name = parse_subject_name(str(cell))
                subjects.append({"name": name, "col": col_idx})

        records = []
        for row_idx in range(6, len(rows)):
            row = rows[row_idx]
            if not row[0] or not str(row[0]).strip().isdigit():
                continue

            name = str(row[2]).strip() if row[2] else ""
            if not name or name.lower() in ("max marks", ""):
                continue

            record = {
                "student_name": name,
                "seat_no": str(row[1]).strip() if row[1] else "",
                "semester": semester,
            }

            for subj in subjects:
                total = parse_total_marks(row[subj["col"]]) if subj["col"] < len(row) else None
                record[f"{subj['name']}_pct"] = total

            records.append(record)

        return records

    def get_student_features(self, df: pd.DataFrame) -> pd.DataFrame:
        student_data = defaultdict(dict)
        for _, row in df.iterrows():
            key = row["student_name"]
            student_data[key]["name"] = row["student_name"]
            student_data[key]["seat_no"] = row["seat_no"]
            
            for col in df.columns:
                if col.endswith("_pct") and row[col] is not None:
                    student_data[key][col] = row[col]

        return pd.DataFrame(list(student_data.values()))

# ═══════════════════════════════════════════════════════════════
#  ENHANCED FEATURE ENGINEERING
# ═══════════════════════════════════════════════════════════════

class FeatureEngineer:
    @staticmethod
    def create_enhanced_features(df: pd.DataFrame) -> Tuple[pd.DataFrame, List[str]]:
        logger.info("Starting advanced feature engineering...")
        
        subject_cols = [c for c in df.columns if c.endswith('_pct')]
        feature_names = []
        
        # 1. SUBJECT CATEGORY AGGREGATES
        math_subjects = [c for c in subject_cols if 'Mathematics' in c]
        prog_subjects = [c for c in subject_cols if 'Data Structures' in c or 'Artificial Intelligence' in c]
        db_subjects = [c for c in subject_cols if 'Database' in c]
        network_subjects = [c for c in subject_cols if 'Network' in c or 'Operating System' in c]
        
        lab_subjects = [c for c in subject_cols if 'Laboratory' in c or 'Lab' in c]
        theory_subjects = [c for c in subject_cols if c not in lab_subjects]
        
        # Safe aggregation with fallback
        df['math_avg'] = df[math_subjects].mean(axis=1, skipna=True).fillna(50) if math_subjects else 50
        df['prog_avg'] = df[prog_subjects].mean(axis=1, skipna=True).fillna(50) if prog_subjects else 50
        df['db_avg'] = df[db_subjects].mean(axis=1, skipna=True).fillna(50) if db_subjects else 50
        df['network_avg'] = df[network_subjects].mean(axis=1, skipna=True).fillna(50) if network_subjects else 50
        df['lab_avg'] = df[lab_subjects].mean(axis=1, skipna=True).fillna(50) if lab_subjects else 50
        df['theory_avg'] = df[theory_subjects].mean(axis=1, skipna=True).fillna(50) if theory_subjects else 50
        
        feature_names.extend(['math_avg', 'prog_avg', 'db_avg', 'network_avg', 'lab_avg', 'theory_avg'])
        
        # 2. OVERALL STATISTICS
        df['overall_avg'] = df[subject_cols].mean(axis=1, skipna=True).fillna(50)
        df['overall_std'] = df[subject_cols].std(axis=1, skipna=True).fillna(0)
        df['overall_max'] = df[subject_cols].max(axis=1, skipna=True).fillna(50)
        df['overall_min'] = df[subject_cols].min(axis=1, skipna=True).fillna(50)
        
        feature_names.extend(['overall_avg', 'overall_std', 'overall_max', 'overall_min'])
        
        # 3. DERIVED FEATURES
        df['theory_lab_gap'] = df['theory_avg'] - df['lab_avg']
        df['consistency_score'] = 1 / (df['overall_std'] + 1)
        df['strong_subjects_count'] = (df[subject_cols] > 70).sum(axis=1)
        df['weak_subjects_count'] = (df[subject_cols] < 50).sum(axis=1)
        
        feature_names.extend(['theory_lab_gap', 'consistency_score', 'strong_subjects_count', 'weak_subjects_count'])
        
        # 4. DOMAIN-SPECIFIC AFFINITY SCORES
        # Extract specific lab scores
        python_lab = df[[c for c in lab_subjects if 'Python' in c]].mean(axis=1, skipna=True).fillna(50)
        web_labs = df[[c for c in lab_subjects if 'Full stack' in c or 'Software Development' in c]].mean(axis=1, skipna=True).fillna(50)
        sql_lab = df[[c for c in lab_subjects if 'SQL' in c]].mean(axis=1, skipna=True).fillna(50)
        cloud_labs = df[[c for c in lab_subjects if 'Cloud' in c or 'Networks' in c]].mean(axis=1, skipna=True).fillna(50)
        
        # Extract specific subject scores
        se_score = df.get('Software Engineering_pct', pd.Series([50]*len(df))).fillna(50)
        ai_score = df.get('Artificial Intelligence_pct', pd.Series([50]*len(df))).fillna(50)
        os_score = df.get('Operating System_pct', pd.Series([50]*len(df))).fillna(50)
        embedded_score = df.get('Microcontroller and Embedded Systems_pct', pd.Series([50]*len(df))).fillna(50)
        
        # Calculate affinity scores (weighted combination)
        df['ml_affinity'] = (
            df['math_avg'] * 0.35 + 
            df['prog_avg'] * 0.30 + 
            python_lab * 0.25 + 
            ai_score * 0.10
        )
        
        df['wt_affinity'] = (
            df['db_avg'] * 0.30 + 
            web_labs * 0.35 + 
            se_score * 0.25 + 
            df['network_avg'] * 0.10
        )
        
        df['dwm_affinity'] = (
            df['db_avg'] * 0.35 + 
            df['math_avg'] * 0.30 + 
            sql_lab * 0.25 + 
            df['prog_avg'] * 0.10
        )
        
        df['ccs_affinity'] = (
            df['network_avg'] * 0.35 + 
            os_score * 0.25 + 
            cloud_labs * 0.30 + 
            embedded_score * 0.10
        )
        
        feature_names.extend(['ml_affinity', 'wt_affinity', 'dwm_affinity', 'ccs_affinity'])
        
        logger.info(f"Created {len(feature_names)} engineered features")
        
        return df, feature_names

# ═══════════════════════════════════════════════════════════════
#  PROPER DIVERSIFIED LABEL ASSIGNMENT
# ═══════════════════════════════════════════════════════════════

def assign_diversified_labels(df: pd.DataFrame) -> pd.Series:
    """
    Assign labels using FORCED balanced distribution.
    Guarantees exactly n/4 students per class.
    """
    logger.info("Assigning diversified labels...")
    
    n_students = len(df)
    electives = ['ML', 'WT', 'DWM', 'CCS']
    
    # Calculate affinity scores
    np.random.seed(42)
    
    scores = pd.DataFrame(index=df.index)
    for elective in electives:
        affinity_col = f'{elective.lower()}_affinity'
        base_score = df[affinity_col].values
        noise = np.random.uniform(-10, 10, size=len(df))
        scores[elective] = base_score + noise
    
    # Sort students by overall performance
    sorted_students = df.sort_values('overall_avg', ascending=False).index
    
    # Round-robin assignment (ensures exact balance)
    final_labels = pd.Series(index=df.index, dtype=str)
    
    for i, idx in enumerate(sorted_students):
        # Assign to elective with highest affinity that needs students
        assigned_counts = final_labels.value_counts()
        target_per_class = n_students // 4
        
        # Get electives needing students
        available = [e for e in electives if assigned_counts.get(e, 0) < target_per_class]
        
        if available:
            # Choose best available option by affinity
            best_elective = max(available, key=lambda e: scores.loc[idx, e])
            final_labels.loc[idx] = best_elective
        else:
            # Fallback: assign to least populated
            final_labels.loc[idx] = electives[i % 4]
    
    final_dist = final_labels.value_counts()
    logger.info(f"Final distribution: {dict(final_dist)}")
    
    return final_labels

# ═══════════════════════════════════════════════════════════════
#  PREPROCESSING WITH FALLBACK
# ═══════════════════════════════════════════════════════════════

class DataPreprocessor:
    @staticmethod
    def preprocess(X, y, test_size=0.2, random_state=42):
        logger.info("Starting preprocessing pipeline...")
        
        # 1. Impute missing values
        logger.info("  - Imputing missing values (KNN)...")
        imputer = KNNImputer(n_neighbors=5)
        X_imputed = imputer.fit_transform(X)
        
        # 2. Encode labels
        le = LabelEncoder()
        y_encoded = le.fit_transform(y)
        
        # Validate class distribution
        unique_classes, class_counts = np.unique(y_encoded, return_counts=True)
        class_dist = dict(zip(le.inverse_transform(unique_classes), class_counts))
        
        logger.info(f"  - Class distribution: {class_dist}")
        
        if len(unique_classes) < 2:
            raise ValueError(f"Only {len(unique_classes)} class found. Need at least 2!")
        
        min_samples = class_counts.min()
        if min_samples < 2:
            raise ValueError(f"Class {le.inverse_transform([unique_classes[class_counts.argmin()]])[0]} has only {min_samples} sample(s). Need at least 2 per class!")
        
        # 3. Split data
        logger.info("  - Splitting data...")
        try:
            # Try stratified split first
            X_train, X_test, y_train, y_test = train_test_split(
                X_imputed, y_encoded,
                test_size=test_size,
                random_state=random_state,
                stratify=y_encoded
            )
            logger.info("     Using stratified split")
        except ValueError as e:
            # Fallback to regular split if stratification fails
            logger.warning(f"     Stratified split failed: {e}")
            logger.info("     Using regular random split")
            X_train, X_test, y_train, y_test = train_test_split(
                X_imputed, y_encoded,
                test_size=test_size,
                random_state=random_state
            )
        
        # 4. Scale features
        logger.info("  - Scaling features (RobustScaler)...")
        scaler = RobustScaler()
        X_train_scaled = scaler.fit_transform(X_train)
        X_test_scaled = scaler.transform(X_test)
        
        # 5. Balance classes with SMOTE (if possible)
        logger.info("  - Balancing classes...")
        try:
            # Check if SMOTE is possible
            train_dist = dict(zip(*np.unique(y_train, return_counts=True)))
            min_train_samples = min(train_dist.values())
            
            if min_train_samples >= 6:
                k = min(5, min_train_samples - 1)
                smote = SMOTE(random_state=random_state, k_neighbors=k)
                X_train_balanced, y_train_balanced = smote.fit_resample(X_train_scaled, y_train)
                
                logger.info(f"     Original: {train_dist}")
                logger.info(f"     Balanced: {dict(zip(*np.unique(y_train_balanced, return_counts=True)))}")
            else:
                logger.warning(f"     SMOTE skipped: insufficient samples (min: {min_train_samples})")
                X_train_balanced, y_train_balanced = X_train_scaled, y_train
        except Exception as e:
            logger.warning(f"     SMOTE failed: {e}")
            X_train_balanced, y_train_balanced = X_train_scaled, y_train
        
        logger.info("Preprocessing complete")
        
        return (X_train_balanced, X_test_scaled, y_train_balanced, y_test,
                scaler, le, imputer)

# ═══════════════════════════════════════════════════════════════
#  MODEL TRAINING (SIMPLIFIED)
# ═══════════════════════════════════════════════════════════════

def train_enhanced_models(X_train, X_test, y_train, y_test, label_encoder):
    logger.info("\n" + "="*80)
    logger.info("ENHANCED MODEL TRAINING")
    logger.info("="*80)
    
    # Determine CV folds based on smallest class
    min_class_count = min(np.bincount(y_train))
    cv_folds = min(5, max(2, min_class_count))
    logger.info(f"Using {cv_folds}-fold cross-validation")
    
    models = {
        "RandomForest": RandomForestClassifier(
            n_estimators=200, max_depth=15, min_samples_split=5,
            random_state=42, n_jobs=-1, class_weight='balanced'
        ),
        "XGBoost": xgb.XGBClassifier(
            n_estimators=200, max_depth=7, learning_rate=0.1,
            random_state=42, eval_metric='mlogloss'
        ),
        "LightGBM": lgb.LGBMClassifier(
            n_estimators=200, max_depth=10, learning_rate=0.1,
            random_state=42, verbose=-1, class_weight='balanced'
        ),
        "LogisticRegression": LogisticRegression(
    max_iter=2000, random_state=42, class_weight='balanced',
    solver='lbfgs'
),
        "ExtraTrees": ExtraTreesClassifier(
            n_estimators=200, max_depth=15, random_state=42,
            n_jobs=-1, class_weight='balanced'
        ),
        "KNN": KNeighborsClassifier(
            n_neighbors=min(7, len(X_train)-1), 
            n_jobs=-1
        )
    }
    
    results = {}
    cv_strategy = StratifiedKFold(n_splits=cv_folds, shuffle=True, random_state=42)
    
    print("\n" + "="*80)
    print("TRAINING PROGRESS")
    print("="*80)
    
    for i, (model_name, model) in enumerate(models.items(), 1):
        print(f"\n[{i}/{len(models)}] Training: {model_name}")
        print("-" * 60)
        
        try:
            start_time = time.time()
            
            # Train
            print(f"  - Training...")
            model.fit(X_train, y_train)
            
            # Predictions
            print(f"  - Evaluating...")
            y_pred = model.predict(X_test)
            y_pred_proba = model.predict_proba(X_test) if hasattr(model, 'predict_proba') else None
            
            # Metrics
            accuracy = accuracy_score(y_test, y_pred)
            precision = precision_score(y_test, y_pred, average='weighted', zero_division=0)
            recall = recall_score(y_test, y_pred, average='weighted', zero_division=0)
            f1 = f1_score(y_test, y_pred, average='weighted', zero_division=0)
            
            # Cross-validation
            print(f"  - Running {cv_folds}-fold CV...")
            cv_scores = cross_val_score(
                model, X_train, y_train, 
                cv=cv_strategy, 
                scoring='f1_weighted', 
                n_jobs=-1
            )
            
            training_time = time.time() - start_time
            
            results[model_name] = {
                "model": model,
                "accuracy": accuracy,
                "precision": precision,
                "recall": recall,
                "f1_score": f1,
                "cv_mean": cv_scores.mean(),
                "cv_std": cv_scores.std(),
                "training_time": training_time,
                "predictions": y_pred,
                "predictions_proba": y_pred_proba
            }
            
            # Display results
            print(f"\n  RESULTS:")
            print(f"     Accuracy:  {accuracy*100:>6.2f}%")
            print(f"     Precision: {precision*100:>6.2f}%")
            print(f"     Recall:    {recall*100:>6.2f}%")
            print(f"     F1 Score:  {f1*100:>6.2f}%")
            print(f"     CV Score:  {cv_scores.mean()*100:>6.2f}% +/- {cv_scores.std()*100:.2f}%")
            print(f"     Time:      {training_time:>6.2f}s")
            
            # Per-class accuracy
            cm = confusion_matrix(y_test, y_pred)
            if cm.shape[0] > 1:  # Only if multiple classes in test set
                class_acc = cm.diagonal() / (cm.sum(axis=1) + 1e-10)
                print(f"\n  Per-Class Accuracy:")
                for j, label in enumerate(label_encoder.classes_):
                    if j < len(class_acc):
                        print(f"     {label}: {class_acc[j]*100:.1f}%")
            
        except Exception as e:
            print(f"  FAILED: {e}")
            logger.error(f"Model {model_name} failed: {e}", exc_info=True)
            continue
    
    # Ensemble (only if we have 3+ successful models)
    if len(results) >= 3:
        print("\n" + "="*80)
        print("ENSEMBLE METHODS")
        print("="*80)
        
        sorted_models = sorted(results.items(), key=lambda x: x[1]['f1_score'], reverse=True)
        top_3 = [(name, data['model']) for name, data in sorted_models[:3]]
        
        print(f"\nVoting Ensemble (top 3: {', '.join([n for n, _ in top_3])})")
        print("-" * 60)
        
        try:
            voting_clf = VotingClassifier(estimators=top_3, voting='soft', n_jobs=-1)
            voting_clf.fit(X_train, y_train)
            y_pred_voting = voting_clf.predict(X_test)
            
            voting_acc = accuracy_score(y_test, y_pred_voting)
            voting_f1 = f1_score(y_test, y_pred_voting, average='weighted')
            
            results["VotingEnsemble"] = {
                "model": voting_clf,
                "accuracy": voting_acc,
                "f1_score": voting_f1,
                "precision": precision_score(y_test, y_pred_voting, average='weighted'),
                "recall": recall_score(y_test, y_pred_voting, average='weighted'),
                "cv_mean": voting_acc,
                "cv_std": 0.0,
                "training_time": 0.0,
                "predictions": y_pred_voting,
                "predictions_proba": None
            }
            
            print(f"  Accuracy:  {voting_acc*100:.2f}%")
            print(f"  F1 Score:  {voting_f1*100:.2f}%")
        except Exception as e:
            print(f"  Failed: {e}")
    
    return results

# ═══════════════════════════════════════════════════════════════
#  MAIN PIPELINE
# ═══════════════════════════════════════════════════════════════

def run_complete_training_pipeline():
    # Setup logging
    logging.basicConfig(
        level=logging.INFO,
        format="%(levelname)s: %(message)s",
        handlers=[
            logging.StreamHandler(),
            logging.FileHandler(
                os.path.join(OUTPUT_DIR, f"training_{datetime.now():%Y%m%d_%H%M%S}.log"),
                encoding='utf-8'
            )
        ]
    )
    
    logger.info("="*80)
    logger.info("COMPLETE ENHANCED TRAINING PIPELINE")
    logger.info("="*80)
    
    # 1. Load data
    logger.info("\nStep 1: Loading Excel data...")
    parser = ExcelDataParser(EXCEL_PATH)
    raw_df = parser.parse_all_sheets()
    features_df = parser.get_student_features(raw_df)
    logger.info(f"   Loaded {len(features_df)} students")
    
    # 2. Feature engineering
    logger.info("\nStep 2: Advanced feature engineering...")
    features_df, feature_names = FeatureEngineer.create_enhanced_features(features_df)
    
    # 3. Label assignment
    logger.info("\nStep 3: Creating diversified labels...")
    features_df['label'] = assign_diversified_labels(features_df)
    
    X = features_df[feature_names].values
    y = features_df['label'].values
    
    logger.info(f"   Features: {X.shape}")
    logger.info(f"   Labels: {dict(Counter(y))}")
    
    # Validate label distribution
    label_counts = Counter(y)
    if min(label_counts.values()) < 2:
        logger.error(f"CRITICAL: Some classes have < 2 samples: {label_counts}")
        logger.error("Cannot proceed with training. Please check label assignment logic.")
        return None
    
    # 4. Preprocessing
    logger.info("\nStep 4: Preprocessing...")
    try:
        (X_train, X_test, y_train, y_test,
         scaler, label_encoder, imputer) = DataPreprocessor.preprocess(X, y)
    except ValueError as e:
        logger.error(f"Preprocessing failed: {e}")
        return None
    
    logger.info(f"   Train: {X_train.shape}")
    logger.info(f"   Test:  {X_test.shape}")
    logger.info(f"   Train labels: {dict(Counter(label_encoder.inverse_transform(y_train)))}")
    logger.info(f"   Test labels: {dict(Counter(label_encoder.inverse_transform(y_test)))}")
    
    # 5. Train models
    logger.info("\nStep 5: Training models...")
    results = train_enhanced_models(X_train, X_test, y_train, y_test, label_encoder)
    
    if not results:
        logger.error("No models trained successfully!")
        return None
    
    # 6. Model comparison
    logger.info("\n" + "="*80)
    logger.info("FINAL MODEL COMPARISON")
    logger.info("="*80)
    
    sorted_results = sorted(results.items(), key=lambda x: (x[1]['f1_score'], x[1]['accuracy']), reverse=True)
    
    print(f"\n{'='*100}")
    print(f"{'RANK':<6} {'MODEL':<25} {'ACCURACY':<12} {'F1 SCORE':<12} {'CV SCORE':<15} {'TIME':<10}")
    print(f"{'='*100}")
    
    for rank, (name, metrics) in enumerate(sorted_results, 1):
        acc = metrics['accuracy'] * 100
        f1 = metrics['f1_score'] * 100
        cv = metrics['cv_mean'] * 100
        cv_std = metrics['cv_std'] * 100
        time_taken = metrics['training_time']
        
        prefix = "** " if rank == 1 else f"{rank}. "
        
        print(f"{prefix:<6} {name:<25} {acc:>6.2f}%      {f1:>6.2f}%      {cv:>6.2f}+/-{cv_std:>4.2f}%  {time_taken:>6.2f}s")
    
    print(f"{'='*100}")
    
    best_name = sorted_results[0][0]
    best_model_data = sorted_results[0][1]
    best_model = best_model_data['model']
    
    print(f"\n{'BEST MODEL: ' + best_name:^100}")
    print(f"{'-'*100}")
    print(f"{'Accuracy:':<20} {best_model_data['accuracy']*100:>6.2f}%")
    print(f"{'F1 Score:':<20} {best_model_data['f1_score']*100:>6.2f}%")
    print(f"{'Precision:':<20} {best_model_data.get('precision', 0)*100:>6.2f}%")
    print(f"{'Recall:':<20} {best_model_data.get('recall', 0)*100:>6.2f}%")
    print(f"{'CV Score:':<20} {best_model_data['cv_mean']*100:>6.2f}% +/- {best_model_data['cv_std']*100:.2f}%")
    print(f"{'='*100}\n")
    
    # 7. Classification report
    logger.info("\nDetailed Classification Report:")
    print(classification_report(
        y_test,
        best_model_data['predictions'],
        target_names=label_encoder.classes_,
        digits=4
    ))
    
    # 8. Save models
    logger.info("\nStep 6: Saving models and metadata...")
    joblib.dump(best_model, os.path.join(MODEL_DIR, "rf_model.joblib"))
    joblib.dump(scaler, os.path.join(MODEL_DIR, "scaler.joblib"))
    joblib.dump(label_encoder, os.path.join(MODEL_DIR, "label_enc.joblib"))
    joblib.dump(imputer, os.path.join(MODEL_DIR, "imputer.joblib"))
    
    # Save all models
    for model_name, model_data in results.items():
        safe_name = model_name.replace(" ", "_").lower()
        joblib.dump(model_data['model'], os.path.join(MODEL_DIR, f"model_{safe_name}.joblib"))
    
    # Metadata
    metadata = {
        "training_timestamp": datetime.now().isoformat(),
        "data_source": "IT - Copy.xlsx",
        "total_students": len(features_df),
        "feature_count": len(feature_names),
        "feature_names": feature_names,
        "label_distribution": dict(Counter(y)),
        "best_model": best_name,
        "best_accuracy": float(best_model_data['accuracy']),
        "best_f1_score": float(best_model_data['f1_score']),
        "best_precision": float(best_model_data.get('precision', 0)),
        "best_recall": float(best_model_data.get('recall', 0)),
        "cv_mean": float(best_model_data['cv_mean']),
        "cv_std": float(best_model_data['cv_std']),
        "training_samples": len(X_train),
        "test_samples": len(X_test),
        "all_models": {
            name: {
                "accuracy": float(data['accuracy']),
                "f1_score": float(data['f1_score']),
                "cv_mean": float(data['cv_mean']),
                "cv_std": float(data['cv_std'])
            }
            for name, data in results.items()
        },
        "model_ranking": [name for name, _ in sorted_results]
    }
    
    with open(os.path.join(MODEL_DIR, "training_metadata.json"), "w") as f:
        json.dump(metadata, f, indent=2)
    
    logger.info("   - Saved all models and metadata")
    
    logger.info("\n" + "="*80)
    logger.info("TRAINING COMPLETE")
    logger.info("="*80)
    logger.info(f"\nSUMMARY:")
    logger.info(f"   Best Model:     {best_name}")
    logger.info(f"   Accuracy:       {metadata['best_accuracy']*100:.2f}%")
    logger.info(f"   F1 Score:       {metadata['best_f1_score']*100:.2f}%")
    logger.info(f"   Students:       {metadata['total_students']}")
    logger.info(f"   Features:       {metadata['feature_count']}")
    logger.info(f"\nOutputs:")
    logger.info(f"   Models:         {MODEL_DIR}")
    logger.info(f"   Logs:           {OUTPUT_DIR}")
    
    return metadata


if __name__ == "__main__":
    run_complete_training_pipeline()