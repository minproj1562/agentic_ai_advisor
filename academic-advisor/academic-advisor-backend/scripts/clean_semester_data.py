#!/usr/bin/env python3
"""
Excel Sheet Splitter — Split multi-sheet Excel into separate files
Usage:
    python clean_semester_data.py input.xlsx output_folder/
"""

import sys
import os
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
import re

def extract_semester_from_sheet_name(sheet_name: str):
    """Extract semester number from sheet name"""
    s = sheet_name.strip().upper()
    
    # Roman numerals
    roman_map = {'I':1, 'II':2, 'III':3, 'IV':4, 'V':5, 'VI':6, 'VII':7, 'VIII':8}
    for roman, num in roman_map.items():
        if f"-{roman}-" in s or f"SEM-{roman}" in s or f" {roman} " in s:
            return num
    
    # Numeric patterns
    match = re.search(r'SEM(?:ESTER)?\s*(\d)', s)
    if match:
        return int(match.group(1))
    
    match = re.search(r'-(\d+)-', s)
    if match:
        return int(match.group(1))
    
    match = re.search(r'\b(\d)\b', s)
    if match:
        return int(match.group(1))
    
    return None


def split_excel_sheets(input_file: str, output_folder: str = None):
    """
    Split a multi-sheet Excel file into separate files.
    """
    input_path = Path(input_file)
    
    if not input_path.exists():
        print(f"❌ Error: File not found: {input_file}")
        return
    
    if not input_path.suffix.lower() in ['.xlsx', '.xls']:
        print(f"❌ Error: Not an Excel file: {input_file}")
        return
    
    # Output folder
    if output_folder:
        output_path = Path(output_folder)
    else:
        output_path = input_path.parent / f"{input_path.stem}_split"
    
    output_path.mkdir(parents=True, exist_ok=True)
    
    print(f"📂 Input file: {input_path}")
    print(f"📁 Output folder: {output_path}")
    print()
    
    # Load workbook
    try:
        wb = load_workbook(input_path, data_only=True)
    except Exception as e:
        print(f"❌ Error loading Excel file: {e}")
        return
    
    sheet_names = wb.sheetnames
    print(f"📊 Found {len(sheet_names)} sheets:")
    for i, name in enumerate(sheet_names, 1):
        sem = extract_semester_from_sheet_name(name)
        print(f"   {i}. {name} {'→ Semester ' + str(sem) if sem else ''}")
    print()
    
    # Split each sheet
    success_count = 0
    for sheet_name in sheet_names:
        try:
            # Read the specific sheet
            df = pd.read_excel(
                input_path,
                sheet_name=sheet_name,
                header=None,
                dtype=object,
            )
            
            # Determine output filename
            sem = extract_semester_from_sheet_name(sheet_name)
            # 🔥 FIX: avoid backslash inside f-string
            safe_base = sheet_name.replace("/", "_").replace("\\", "_")
            if sem:
                safe_name = f"Semester_{sem}_{safe_base}"
            else:
                safe_name = safe_base
            
            output_file = output_path / f"{safe_name}.xlsx"
            
            # Create new workbook with single sheet
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            new_wb = Workbook()
            new_ws = new_wb.active
            new_ws.title = sheet_name[:31]
            
            # Copy data
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 1):
                for c_idx, value in enumerate(row, 1):
                    new_ws.cell(row=r_idx, column=c_idx, value=value)
            
            new_wb.save(output_file)
            print(f"✅ Saved: {output_file.name}")
            success_count += 1
            
        except Exception as e:
            print(f"❌ Error processing '{sheet_name}': {e}")
    
    print()
    print(f"✨ Done! Split {success_count}/{len(sheet_names)} sheets successfully.")
    print(f"📁 Output folder: {output_path.absolute()}")
    print()
    print("📋 Next steps:")
    print("   1. Review the split files")
    print("   2. Upload each file separately in the admin portal:")
    print("      Admin → Bulk Marks Upload → Select semester → Upload file")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python clean_semester_data.py <input_file.xlsx> [output_folder]")
        print()
        print("Example:")
        print("  python clean_semester_data.py IT_marks.xlsx")
        print("  python clean_semester_data.py IT_marks.xlsx ./split_output")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_folder = sys.argv[2] if len(sys.argv) > 2 else None
    
    split_excel_sheets(input_file, output_folder)