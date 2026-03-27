#academic-advisor/academic-advisor-backend/scripts/generate_marks_template_with_students.py
"""
Generate marks templates with student data pre-filled (including seat numbers)
Supports both roll number and seat number identification

Usage:
  python scripts/generate_marks_template_with_students.py -s 5 -b IT
  python scripts/generate_marks_template_with_students.py -s 5 -b IT --seat-numbers
  python scripts/generate_marks_template_with_students.py --all -b IT -a 2022
"""

import argparse
import asyncio
import sys
import os
from datetime import datetime
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from typing import List, Optional, Dict, Any

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.models.student_profile import StudentProfile, SeatNumberRecord
from app.models.pending_marks import PendingStudentMarks
from app.services.bulk_marks_service import bulk_marks_service
from app.core.curriculum import get_semester_subjects
from motor.motor_asyncio import AsyncIOMotorClient
from beanie import init_beanie
from dotenv import load_dotenv

load_dotenv()

# Styles
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

async def init_db():
    """Initialize database connection"""
    client = AsyncIOMotorClient(os.getenv("MONGODB_URL", "mongodb://localhost:27017"))
    db = client[os.getenv("DB_NAME", "academic_advisor")]
    await init_beanie(
        database=db,
        document_models=[StudentProfile, PendingStudentMarks]
    )

async def fetch_students_with_details(
    branch: str, 
    admission_year: Optional[int] = None,
    semester: Optional[int] = None
) -> List[Dict[str, Any]]:
    """Fetch students with their seat numbers for the specified semester"""
    query = {}
    
    if branch:
        query["branch"] = {"$regex": f"^{branch}$", "$options": "i"}
    if admission_year:
        query["admission_year"] = admission_year
    
    students = await StudentProfile.find(query).sort("roll_number").to_list()
    
    student_data = []
    for student in students:
        # Find seat number for the specified semester
        seat_number = None
        
        # First check current seat number if it's current semester
        if semester and student.current_semester == semester:
            seat_number = student.current_seat_number
        
        # Otherwise check history
        if not seat_number and semester:
            for seat_record in student.seat_number_history:
                if seat_record.semester == semester:
                    seat_number = seat_record.seat_number
                    break
        
        # If no semester specified or no history, use current
        if not seat_number:
            seat_number = student.current_seat_number
        
        student_data.append({
            "roll_number": student.roll_number,
            "seat_number": seat_number or "",
            "name": student.name,
            "branch": student.branch,
            "admission_year": student.admission_year,
            "current_semester": student.current_semester,
            "cgpa": student.cgpa,
            "email": student.email
        })
    
    return student_data

async def create_marks_template_with_identifiers(
    semester: int,
    branch: str,
    admission_year: int,
    students: List[Dict[str, Any]],
    output_file: str,
    use_seat_numbers: bool = False,
    academic_year: str = "2024-25"
):
    """
    Create Excel template with both roll numbers and seat numbers
    """
    
    # Generate the base template
    print(f"Generating template for Semester {semester}, {branch} branch...")
    template_buffer = bulk_marks_service.generate_template(
        semester=semester,
        branch=branch,
        academic_year=academic_year,
        admission_year=admission_year,
        elective_choices=None
    )
    
    # Load the generated template
    wb = load_workbook(template_buffer)
    ws = wb["Marks Data"]
    
    # Find where student data starts (typically row 7)
    DATA_START_ROW = 7
    
    # Read column mapping if available
    col_map_json = None
    try:
        marker_cell = ws.cell(row=100, column=1).value
        if marker_cell == "__COLUMN_MAP__":
            col_map_json = ws.cell(row=100, column=2).value
            col_map = json.loads(col_map_json) if col_map_json else []
    except:
        print("Warning: Could not find column map in template")
        col_map = []
    
    # Modify the header to clarify Roll Number / Seat Number usage
    seat_no_cell = ws.cell(row=4, column=2)  # The "Seat No" header
    if use_seat_numbers:
        seat_no_cell.value = "Seat No\n(6-digit)"
        ws.merge_cells(start_row=4, start_column=2, end_row=5, end_column=2)
    else:
        seat_no_cell.value = "Roll Number /\nSeat No"
        ws.merge_cells(start_row=4, start_column=2, end_row=5, end_column=2)
    
    # Apply styling to the modified header
    seat_no_cell.font = Font(bold=True, color="FFFFFF", size=10)
    seat_no_cell.fill = PatternFill("solid", fgColor="2F5496")
    seat_no_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    seat_no_cell.border = THIN_BORDER
    
    # Add an additional column for alternate identifier if needed
    if not use_seat_numbers:
        # Insert new column after Name column (column 3)
        ws.insert_cols(4)
        
        # Add header for seat number column
        seat_header = ws.cell(row=4, column=4, value="Seat Number\n(Optional)")
        seat_header.font = Font(bold=True, color="FFFFFF", size=10)
        seat_header.fill = PatternFill("solid", fgColor="2F5496")
        seat_header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        seat_header.border = THIN_BORDER
        ws.merge_cells(start_row=4, start_column=4, end_row=5, end_column=4)
        
        # Update column widths
        ws.column_dimensions['D'].width = 15
    
    # Fill in student data
    print(f"Adding {len(students)} students to template...")
    
    for idx, student in enumerate(students):
        row = DATA_START_ROW + idx
        
        # Sr. No
        ws.cell(row=row, column=1, value=idx + 1)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=1).border = THIN_BORDER
        
        # Primary Identifier (Roll Number or Seat Number based on flag)
        primary_id = student['seat_number'] if use_seat_numbers and student['seat_number'] else student['roll_number']
        ws.cell(row=row, column=2, value=primary_id)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2).font = Font(bold=True)
        ws.cell(row=row, column=2).border = THIN_BORDER
        
        # Name
        ws.cell(row=row, column=3, value=student['name'])
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="left")
        ws.cell(row=row, column=3).border = THIN_BORDER
        
        # Additional identifier column (if not using seat numbers as primary)
        if not use_seat_numbers:
            ws.cell(row=row, column=4, value=student['seat_number'] or "")
            ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=4).border = THIN_BORDER
            ws.cell(row=row, column=4).font = Font(italic=True, color="666666")
        
        # Apply borders to all cells in the row
        last_col = ws.max_column
        for col in range(1, last_col + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER
    
    # Add metadata about the student list
    metadata_row = 101
    ws.cell(row=metadata_row, column=1, value="__STUDENTS_META__")
    ws.cell(row=metadata_row, column=2, value=json.dumps({
        "total_students": len(students),
        "branch": branch,
        "admission_year": admission_year,
        "semester": semester,
        "prefilled": True,
        "identifier_type": "seat_number" if use_seat_numbers else "roll_number",
        "generated_at": datetime.now().isoformat()
    }))
    ws.row_dimensions[metadata_row].hidden = True
    
    # Add instructions in the Subject Info sheet
    if "Subject Info" in wb.sheetnames:
        info_sheet = wb["Subject Info"]
        last_row = info_sheet.max_row + 2
        info_sheet.cell(row=last_row, column=1, value="Student Identification:")
        info_sheet.cell(row=last_row, column=1).font = Font(bold=True)
        
        instructions = [
            f"- Primary identifier used: {'Seat Numbers' if use_seat_numbers else 'Roll Numbers'}",
            f"- Total students pre-filled: {len(students)}",
            f"- Branch: {branch}, Semester: {semester}",
            "- You can use either Roll Number OR Seat Number for matching",
            "- System will automatically match students by either identifier"
        ]
        
        for i, instruction in enumerate(instructions):
            info_sheet.cell(row=last_row + 1 + i, column=1, value=instruction)
    
    # Save the file
    wb.save(output_file)
    print(f"\n✅ Excel file created: {output_file}")
    print(f"   - Students: {len(students)}")
    print(f"   - Primary ID: {'Seat Numbers' if use_seat_numbers else 'Roll Numbers'}")
    print(f"   - Format: University marks template (compatible with bulk upload)")

async def generate_summary_report(students: List[Dict[str, Any]], output_file: str):
    """Generate a summary report of students with missing seat numbers"""
    
    missing_seat = [s for s in students if not s['seat_number']]
    
    if missing_seat:
        print(f"\n⚠️  Warning: {len(missing_seat)} students have no seat number:")
        with open(output_file, 'w') as f:
            f.write(f"Students Missing Seat Numbers - Generated {datetime.now()}\n")
            f.write("=" * 60 + "\n\n")
            
            for student in missing_seat:
                f.write(f"Roll Number: {student['roll_number']}\n")
                f.write(f"Name: {student['name']}\n")
                f.write(f"Branch: {student['branch']}\n")
                f.write(f"Current Semester: {student['current_semester']}\n")
                f.write("-" * 40 + "\n")
        
        print(f"   Report saved to: {output_file}")

async def main():
    parser = argparse.ArgumentParser(
        description="Generate marks template with student data (roll numbers and seat numbers)"
    )
    parser.add_argument("-s", "--semester", type=int, required=True, help="Semester number (1-8)")
    parser.add_argument("-b", "--branch", required=True, help="Branch code (e.g., IT, COMP)")
    parser.add_argument("-a", "--admission", type=int, help="Admission year filter")
    parser.add_argument("-y", "--year", default="2024-25", help="Academic year (default: 2024-25)")
    parser.add_argument("--seat-numbers", action="store_true", 
                       help="Use seat numbers as primary identifier instead of roll numbers")
    parser.add_argument("--all", action="store_true", 
                       help="Generate templates for all semesters (1-8)")
    parser.add_argument("-o", "--output", help="Output filename/directory")
    parser.add_argument("--report", action="store_true", 
                       help="Generate report of students with missing data")
    
    args = parser.parse_args()
    
    # Initialize database
    await init_db()
    
    if args.all:
        # Generate for all semesters
        output_dir = args.output or "."
        os.makedirs(output_dir, exist_ok=True)
        
        for sem in range(1, 9):
            print(f"\n{'='*60}")
            print(f"Processing Semester {sem}...")
            print('='*60)
            
            # Fetch students
            students = await fetch_students_with_details(
                branch=args.branch,
                admission_year=args.admission,
                semester=sem
            )
            
            if not students:
                print(f"❌ No students found for semester {sem}")
                continue
            
            # Generate filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            id_type = "seat" if args.seat_numbers else "roll"
            filename = f"marks_template_sem{sem}_{args.branch}_{id_type}_{timestamp}.xlsx"
            output_file = os.path.join(output_dir, filename)
            
            # Create template
            await create_marks_template_with_identifiers(
                semester=sem,
                branch=args.branch,
                admission_year=args.admission or 2022,
                students=students,
                output_file=output_file,
                use_seat_numbers=args.seat_numbers,
                academic_year=args.year
            )
            
            # Generate report if requested
            if args.report:
                report_file = os.path.join(output_dir, f"missing_data_sem{sem}_{args.branch}.txt")
                await generate_summary_report(students, report_file)
        
        print(f"\n{'='*60}")
        print(f"✅ All templates generated in: {os.path.abspath(output_dir)}")
        
    else:
        # Single semester
        print(f"Fetching students from {args.branch} branch, semester {args.semester}...")
        students = await fetch_students_with_details(
            branch=args.branch,
            admission_year=args.admission,
            semester=args.semester
        )
        
        if not students:
            print("❌ No students found with the given criteria")
            return
        
        print(f"✅ Found {len(students)} students")
        
        # Check seat number availability
        with_seat = len([s for s in students if s['seat_number']])
        print(f"   - With seat numbers: {with_seat}")
        print(f"   - Without seat numbers: {len(students) - with_seat}")
        
        # Generate output filename
        if args.output:
            output_file = args.output
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            id_type = "seat" if args.seat_numbers else "roll"
            output_file = f"marks_template_sem{args.semester}_{args.branch}_{id_type}_{timestamp}.xlsx"
        
        # Create Excel file
        await create_marks_template_with_identifiers(
            semester=args.semester,
            branch=args.branch,
            admission_year=args.admission or 2022,
            students=students,
            output_file=output_file,
            use_seat_numbers=args.seat_numbers,
            academic_year=args.year
        )
        
        # Generate report if requested
        if args.report:
            report_file = f"missing_data_{args.branch}_sem{args.semester}.txt"
            await generate_summary_report(students, report_file)
    
    print("\n📋 Next steps:")
    print("1. Open the Excel file(s)")
    print("2. Fill in the marks for each student")
    print("3. Save the file")
    print("4. Upload through Admin Panel > Bulk Marks Upload")
    print("\n💡 Tips:")
    print("- The system will match students by EITHER roll number OR seat number")
    print("- Students without profiles will have marks saved as 'pending'")
    print("- Pending marks auto-sync when students create their profiles")

if __name__ == "__main__":
    asyncio.run(main())