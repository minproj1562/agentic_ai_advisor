# academic-advisor-backend/scripts/generate_full_marks_export.py
"""
Comprehensive Marks Generator & Exporter
=========================================
Uses bulk_marks_service.generate_template() to produce EXACT university-format
templates (with __COLUMN_MAP__), then fills them with generated/existing marks.

Files are fully compatible with the bulk upload system.

Usage:
  python -m scripts.generate_full_marks_export
  python -m scripts.generate_full_marks_export --export-only
  python -m scripts.generate_full_marks_export --seed-only
  python -m scripts.generate_full_marks_export --branch IT
  python -m scripts.generate_full_marks_export --admission-year 2022
  python -m scripts.generate_full_marks_export --dry-run
  python -m scripts.generate_full_marks_export --overwrite
"""

import asyncio
import argparse
import io
import json
import os
import sys
import random
import logging
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from motor.motor_asyncio import AsyncIOMotorClient
from beanie import init_beanie
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════
# 1.  PERFORMANCE PROFILES
# ══════════════════════════════════════════════════════════

@dataclass
class PerformanceProfile:
    name: str
    internal_mean: float
    internal_std: float
    external_mean: float
    external_std: float
    fail_probability: float
    practical_bonus: float
    weight: float


PERFORMANCE_PROFILES = [
    PerformanceProfile("outstanding",    92, 4,  88, 5,  0.00, 5, 0.05),
    PerformanceProfile("excellent",      84, 6,  78, 8,  0.00, 5, 0.10),
    PerformanceProfile("very_good",      76, 7,  70, 9,  0.02, 4, 0.15),
    PerformanceProfile("good",           68, 8,  62, 10, 0.05, 3, 0.20),
    PerformanceProfile("above_average",  62, 9,  56, 11, 0.08, 3, 0.15),
    PerformanceProfile("average",        55, 10, 50, 12, 0.12, 2, 0.15),
    PerformanceProfile("below_average",  48, 10, 42, 13, 0.20, 2, 0.10),
    PerformanceProfile("struggling",     42, 12, 36, 14, 0.35, 2, 0.07),
    PerformanceProfile("at_risk",        35, 12, 30, 15, 0.50, 1, 0.03),
]


def assign_performance_profile() -> PerformanceProfile:
    weights = [p.weight for p in PERFORMANCE_PROFILES]
    return random.choices(PERFORMANCE_PROFILES, weights=weights, k=1)[0]


# ══════════════════════════════════════════════════════════
# 2.  MARKS GENERATION
# ══════════════════════════════════════════════════════════

def generate_marks_for_subject(
    profile: PerformanceProfile,
    internal_max: float,
    external_max: float,
    is_practical: bool = False,
    is_elective: bool = False,
    semester_drift: float = 0.0,
) -> Tuple[float, float]:
    int_pct = random.gauss(profile.internal_mean + semester_drift, profile.internal_std)
    ext_pct = random.gauss(profile.external_mean + semester_drift, profile.external_std)

    if is_practical:
        int_pct += profile.practical_bonus
        ext_pct += profile.practical_bonus
    if is_elective:
        int_pct += random.uniform(1, 4)
        ext_pct += random.uniform(1, 4)

    if random.random() < profile.fail_probability:
        ext_pct = random.uniform(10, 38)
        if random.random() < 0.4:
            int_pct = random.uniform(15, 40)

    int_pct = max(0, min(100, int_pct))
    ext_pct = max(0, min(100, ext_pct))

    internal = round((int_pct / 100.0) * internal_max, 1)
    external = round((ext_pct / 100.0) * external_max, 1)
    internal = max(0, min(internal, internal_max))
    external = max(0, min(external, external_max))

    return internal, external


def calculate_grade(total: float, max_marks: float) -> Dict[str, Any]:
    if max_marks <= 0:
        max_marks = 100.0
    pct = (total / max_marks) * 100.0
    if pct >= 90: return {"grade": "O",  "points": 10.0}
    if pct >= 80: return {"grade": "A+", "points":  9.0}
    if pct >= 70: return {"grade": "A",  "points":  8.0}
    if pct >= 60: return {"grade": "B+", "points":  7.0}
    if pct >= 50: return {"grade": "B",  "points":  6.0}
    if pct >= 45: return {"grade": "C",  "points":  5.0}
    if pct >= 40: return {"grade": "P",  "points":  4.0}
    return {"grade": "F", "points": 0.0}


def get_academic_year(semester: int, admission_year: int) -> str:
    year_offset = (semester - 1) // 2
    start_year = admission_year + year_offset
    return f"{start_year}-{str(start_year + 1)[2:]}"


def get_max_semester_for_student(admission_year: int) -> int:
    now = datetime.now()
    if now.month >= 7:
        current_sem = (now.year - admission_year) * 2 + 1
    else:
        current_sem = (now.year - admission_year) * 2
    return max(1, min(current_sem, 8))


# ══════════════════════════════════════════════════════════
# 3.  COMPONENT SPLITTING (mirrors bulk_marks_service)
# ══════════════════════════════════════════════════════════

def split_marks_to_components(
    internal: float, external: float,
    course_type: str, is_practical: bool,
    internal_max: float, external_max: float,
) -> Dict[str, float]:
    """
    Split internal/external → university component marks.
    Exact reverse of components_to_subject_score() in bulk_marks_service.
    """
    ct = course_type

    if ct in ("MNP", "MJP", "INT"):
        return {"TW": internal}

    if ct == "SBL":
        return {"TW": internal, "PR": external}

    if ct == "LBC" or is_practical:
        return {"IA": internal, "PR": external}

    # Theory: CA = internal, MSE+ESE = external
    ca = internal
    if external_max == 80:
        mse_max, ese_max = 30, 50
    elif external_max == 0:
        return {"CA": ca}
    else:
        mse_max = round(external_max * 0.375)
        ese_max = external_max - mse_max

    total_ext_max = mse_max + ese_max
    if total_ext_max > 0 and external > 0:
        mse = round(external * (mse_max / total_ext_max), 1)
        ese = round(external - mse, 1)
    else:
        mse, ese = 0.0, 0.0

    return {"CA": ca, "MSE": mse, "ESE": ese}


# ══════════════════════════════════════════════════════════
# 4.  DATABASE INIT
# ══════════════════════════════════════════════════════════

async def init_database():
    from app.core.config import settings
    from app.models.student_profile import StudentProfile
    from app.models.pending_marks import PendingStudentMarks

    client = AsyncIOMotorClient(settings.MONGODB_URL)
    db = client[settings.DATABASE_NAME]
    await init_beanie(
        database=db,
        document_models=[StudentProfile, PendingStudentMarks],
    )
    print(f"✅ Connected to MongoDB: {settings.DATABASE_NAME}")
    return client


# ══════════════════════════════════════════════════════════
# 5.  SEED MARKS INTO MONGODB
# ══════════════════════════════════════════════════════════

async def seed_marks_for_all_students(
    branch_filter: Optional[str] = None,
    admission_year_filter: Optional[int] = None,
    overwrite: bool = False,
    dry_run: bool = False,
) -> Dict[str, Any]:
    from app.models.student_profile import StudentProfile, SemesterRecord, SubjectScore
    from app.core.curriculum import get_semester_subjects

    query: Dict[str, Any] = {}
    if branch_filter:
        query["branch"] = {"$regex": f"^{branch_filter}$", "$options": "i"}
    if admission_year_filter:
        query["admission_year"] = admission_year_filter

    students = await StudentProfile.find(query).sort("roll_number").to_list()

    if not students:
        print("❌ No students found")
        return {"total_students": 0, "updated": 0, "skipped": 0,
                "semesters_created": 0, "semesters_skipped_existing": 0,
                "profile_distribution": {}, "grade_distribution": {},
                "by_admission_year": {}}

    print(f"\n👥 Found {len(students)} students")
    if not overwrite:
        has_marks = sum(1 for s in students if any(sr.subjects for sr in s.semester_records))
        print(f"   {has_marks} already have marks (use --overwrite to regenerate)")

    stats = {
        "total_students": len(students),
        "updated": 0,
        "skipped": 0,
        "semesters_created": 0,
        "semesters_skipped_existing": 0,
        "profile_distribution": {},
        "grade_distribution": {
            "O": 0, "A+": 0, "A": 0, "B+": 0,
            "B": 0, "C": 0, "P": 0, "F": 0,
        },
        "by_admission_year": {},
    }

    for idx, student in enumerate(students):
        admission_year = student.admission_year
        max_sem = get_max_semester_for_student(admission_year)
        semesters_to_generate = list(range(1, max_sem + 1))

        perf_profile = assign_performance_profile()
        stats["profile_distribution"][perf_profile.name] = (
            stats["profile_distribution"].get(perf_profile.name, 0) + 1
        )

        yr_key = str(admission_year)
        if yr_key not in stats["by_admission_year"]:
            stats["by_admission_year"][yr_key] = {
                "students": 0, "semesters_total": 0, "cgpa_sum": 0.0,
            }
        stats["by_admission_year"][yr_key]["students"] += 1

        student_updated = False
        semester_drift = 0.0

        for sem_num in semesters_to_generate:
            existing_sem = next(
                (sr for sr in student.semester_records
                 if sr.semester_number == sem_num),
                None,
            )

            # Skip if already has subjects and not overwriting
            if existing_sem and existing_sem.subjects and not overwrite:
                stats["semesters_skipped_existing"] += 1
                continue

            subjects = get_semester_subjects(sem_num, admission_year)
            if not subjects:
                continue

            academic_year = get_academic_year(sem_num, admission_year)

            if sem_num > 1:
                semester_drift += random.gauss(0.5, 1.5)
                semester_drift = max(-10, min(10, semester_drift))

            subject_scores: List[SubjectScore] = []
            total_grade_points = 0.0
            total_credits = 0
            credits_earned = 0

            for sub in subjects:
                internal, external = generate_marks_for_subject(
                    profile=perf_profile,
                    internal_max=sub.internal_max,
                    external_max=sub.external_max,
                    is_practical=sub.is_practical,
                    is_elective=sub.is_elective,
                    semester_drift=semester_drift,
                )

                total = round(internal + external, 1)
                max_total = sub.internal_max + sub.external_max
                grade_info = calculate_grade(total, max_total)

                score = SubjectScore(
                    subject_code=sub.subject_code,
                    subject_name=sub.subject_name,
                    credits=sub.credits,
                    internal_marks=internal,
                    external_marks=external,
                    total_marks=total,
                    grade=grade_info["grade"],
                    grade_points=grade_info["points"],
                    is_elective=sub.is_elective,
                    is_practical=sub.is_practical,
                )
                subject_scores.append(score)
                total_credits += sub.credits
                if grade_info["grade"] != "F":
                    total_grade_points += grade_info["points"] * sub.credits
                    credits_earned += sub.credits
                stats["grade_distribution"][grade_info["grade"]] += 1

            sgpa = round(total_grade_points / total_credits, 2) if total_credits > 0 else 0.0

            sem_record = SemesterRecord(
                semester_number=sem_num,
                academic_year=academic_year,
                subjects=subject_scores,
                sgpa=sgpa,
                total_credits=total_credits,
                credits_earned=credits_earned,
                is_complete=True,
                created_at=datetime.now(),
            )

            if existing_sem:
                sem_idx = next(
                    i for i, sr in enumerate(student.semester_records)
                    if sr.semester_number == sem_num
                )
                student.semester_records[sem_idx] = sem_record
            else:
                student.semester_records.append(sem_record)

            student_updated = True
            stats["semesters_created"] += 1
            stats["by_admission_year"][yr_key]["semesters_total"] += 1

        if student_updated:
            student.semester_records.sort(key=lambda x: x.semester_number)

            all_gp = sum(
                sr.sgpa * sr.total_credits
                for sr in student.semester_records
                if sr.is_complete and sr.total_credits > 0
            )
            all_credits = sum(
                sr.total_credits
                for sr in student.semester_records
                if sr.is_complete and sr.total_credits > 0
            )
            all_earned = sum(sr.credits_earned for sr in student.semester_records)

            student.cgpa = round(all_gp / all_credits, 2) if all_credits > 0 else 0.0
            student.total_credits_earned = all_earned
            student.marks_synced_at = datetime.now()
            student.pending_marks_checked = True
            student.last_updated = datetime.now()

            if not dry_run:
                await student.save()

            stats["updated"] += 1
            stats["by_admission_year"][yr_key]["cgpa_sum"] += student.cgpa
        else:
            stats["skipped"] += 1

        if (idx + 1) % 25 == 0 or (idx + 1) == len(students):
            print(
                f"  [{idx + 1}/{len(students)}] "
                f"{perf_profile.name:15s} | "
                f"Sems: {len(student.semester_records)} | "
                f"CGPA: {student.cgpa:.2f} | "
                f"{student.roll_number}"
            )

    for yr_key, yr_data in stats["by_admission_year"].items():
        n = yr_data["students"]
        updated_in_yr = stats["updated"]  # approximate
        if yr_data["cgpa_sum"] > 0 and n > 0:
            yr_data["avg_cgpa"] = round(yr_data["cgpa_sum"] / n, 2)
        else:
            yr_data["avg_cgpa"] = 0.0
        del yr_data["cgpa_sum"]

    return stats


# ══════════════════════════════════════════════════════════
# 6.  EXPORT — one file per (admission_year, branch, semester)
#     Uses bulk_marks_service.generate_template() for format
# ══════════════════════════════════════════════════════════

async def export_all_marks_to_xlsx(
    branch_filter: Optional[str] = None,
    admission_year_filter: Optional[int] = None,
    output_dir: str = "./exported_marks",
) -> List[str]:
    """
    Export ALL marks using the EXACT university template format.

    For each (admission_year, branch, semester):
      1. Call bulk_marks_service.generate_template() → gets proper format
      2. Read __COLUMN_MAP__ from the template
      3. Fill student marks into correct cells
      4. Save to individual .xlsx file
    """
    from app.models.student_profile import StudentProfile
    from app.core.curriculum import get_semester_subjects, get_elective_options
    from app.services.bulk_marks_service import bulk_marks_service

    THIN = Side(style="thin")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    query: Dict[str, Any] = {}
    if branch_filter:
        query["branch"] = {"$regex": f"^{branch_filter}$", "$options": "i"}
    if admission_year_filter:
        query["admission_year"] = admission_year_filter

    students = await StudentProfile.find(query).sort("roll_number").to_list()

    if not students:
        print("❌ No students found")
        return []

    print(f"📊 Exporting marks for {len(students)} students...")

    # Group by (admission_year, branch)
    groups: Dict[Tuple[int, str], List] = {}
    for s in students:
        key = (s.admission_year, s.branch)
        groups.setdefault(key, []).append(s)

    # Find semesters with data per group
    group_semesters: Dict[Tuple[int, str], set] = {}
    for key, grp in groups.items():
        sems = set()
        for s in grp:
            for sr in s.semester_records:
                if sr.subjects:
                    sems.add(sr.semester_number)
        group_semesters[key] = sems

    os.makedirs(output_dir, exist_ok=True)
    output_files: List[str] = []

    for (adm_year, branch), group_students in sorted(groups.items()):
        semesters = sorted(group_semesters.get((adm_year, branch), set()))

        if not semesters:
            print(f"  ⏭️  {branch} {adm_year}: no semesters with marks")
            continue

        for sem_num in semesters:
            subjects = get_semester_subjects(sem_num, adm_year)
            if not subjects:
                continue

            academic_year = get_academic_year(sem_num, adm_year)

            # ── Step 1: Generate official template ──
            try:
                template_buf = bulk_marks_service.generate_template(
                    semester=sem_num,
                    branch=branch,
                    academic_year=academic_year,
                    admission_year=adm_year,
                    elective_choices=None,
                )
            except Exception as e:
                print(f"  ❌ {branch} {adm_year} S{sem_num}: template error: {e}")
                continue

            # ── Step 2: Load and read __COLUMN_MAP__ ──
            wb = load_workbook(template_buf)
            ws = wb["Marks Data"]

            col_map_json = None
            try:
                if ws.cell(row=100, column=1).value == "__COLUMN_MAP__":
                    col_map_json = ws.cell(row=100, column=2).value
            except Exception:
                pass

            if not col_map_json:
                print(f"  ❌ {branch} {adm_year} S{sem_num}: no __COLUMN_MAP__")
                continue

            col_map = json.loads(col_map_json)

            # Build lookup: {subject_code: {components: {key: {col, max}}, elec_code_col, tot_col}}
            subject_columns: Dict[str, Dict[str, Any]] = {}
            for entry in col_map:
                scode = entry["subject_code"]
                if scode not in subject_columns:
                    subject_columns[scode] = {
                        "components": {},
                        "elective_group": entry.get("elective_group", ""),
                        "elec_code_col": None,
                        "tot_col": None,
                    }
                comp = entry["component"]
                if comp == "ELEC_CODE":
                    subject_columns[scode]["elec_code_col"] = entry["col"]
                elif comp == "TOT":
                    subject_columns[scode]["tot_col"] = entry["col"]
                else:
                    subject_columns[scode]["components"][comp] = {
                        "col": entry["col"],
                        "max": entry["max"],
                    }

            # Find data start row
            DATA_START = 7
            for r in range(4, 20):
                v = ws.cell(row=r, column=1).value
                if v is not None and str(v).strip().isdigit():
                    DATA_START = r
                    break

            # Find summary columns (after all subject columns)
            all_col_nums = [e["col"] for e in col_map]
            max_subject_col = max(all_col_nums) if all_col_nums else 3
            summary_total_col = max_subject_col + 1
            summary_sgpi_col = max_subject_col + 2
            summary_result_col = max_subject_col + 3

            # Build elective code lookup
            elec_code_lookup: Dict[str, Dict[str, str]] = {}
            for sub in subjects:
                if sub.is_elective and sub.elective_group:
                    for opt in get_elective_options(sub.elective_group):
                        elec_code_lookup[opt["code"].upper()] = {
                            "template_code": sub.subject_code,
                            "name": opt["name"],
                            "code": opt["code"],
                        }

            # ── Step 3: Fill student data ──
            sorted_students = sorted(group_students, key=lambda s: s.roll_number)
            students_filled = 0
            students_no_marks = 0

            for idx, student in enumerate(sorted_students):
                row = DATA_START + idx

                # Fixed columns
                ws.cell(row=row, column=1, value=idx + 1)
                ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=1).border = BORDER

                roll_cell = ws.cell(row=row, column=2, value=student.roll_number)
                roll_cell.alignment = Alignment(horizontal="center")
                roll_cell.font = Font(bold=True)
                roll_cell.border = BORDER

                name_cell = ws.cell(row=row, column=3, value=student.name)
                name_cell.alignment = Alignment(horizontal="left")
                name_cell.border = BORDER

                # Find semester record
                sem_record = next(
                    (sr for sr in student.semester_records
                     if sr.semester_number == sem_num),
                    None,
                )

                if not sem_record or not sem_record.subjects:
                    students_no_marks += 1
                    for c in range(1, summary_result_col + 1):
                        ws.cell(row=row, column=c).border = BORDER
                    continue

                students_filled += 1

                # Green highlight for rows with marks
                for c in range(1, 4):
                    ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor="E2EFDA")

                total_marks_sum = 0.0

                for subj_score in sem_record.subjects:
                    sc = subj_score.subject_code

                    # Find column info: direct match
                    col_info = subject_columns.get(sc)

                    # Try elective lookup
                    if not col_info and sc.upper() in elec_code_lookup:
                        template_code = elec_code_lookup[sc.upper()]["template_code"]
                        col_info = subject_columns.get(template_code)

                    # Try by subject name
                    if not col_info:
                        for tmpl_code, tmpl_info in subject_columns.items():
                            tmpl_sub = next(
                                (s for s in subjects if s.subject_code == tmpl_code),
                                None,
                            )
                            if (tmpl_sub and
                                tmpl_sub.subject_name.lower() == subj_score.subject_name.lower()):
                                col_info = tmpl_info
                                break

                    if not col_info:
                        continue

                    internal = subj_score.internal_marks
                    external = subj_score.external_marks
                    total = subj_score.total_marks
                    total_marks_sum += total

                    # Get curriculum subject for this code
                    curr_sub = next(
                        (s for s in subjects if s.subject_code == sc), None
                    )
                    # Fallback: find via template code
                    if not curr_sub:
                        for tmpl_code in subject_columns:
                            if subject_columns[tmpl_code] is col_info:
                                curr_sub = next(
                                    (s for s in subjects if s.subject_code == tmpl_code),
                                    None,
                                )
                                break

                    if curr_sub:
                        ct = curr_sub.course_type
                        ip = curr_sub.is_practical
                        im = curr_sub.internal_max
                        em = curr_sub.external_max
                    else:
                        ct, ip, im, em = "PCC", False, 20.0, 80.0

                    # Split into components
                    comp_marks = split_marks_to_components(
                        internal, external, ct, ip, im, em
                    )

                    # Fill elective code column
                    if col_info.get("elec_code_col"):
                        elec_info = elec_code_lookup.get(sc.upper())
                        if elec_info:
                            ec_cell = ws.cell(
                                row=row,
                                column=col_info["elec_code_col"],
                                value=elec_info["code"],
                            )
                            ec_cell.alignment = Alignment(horizontal="center")
                            ec_cell.border = BORDER

                    # Fill component columns
                    for comp_key, mark_value in comp_marks.items():
                        comp_info = col_info["components"].get(comp_key)
                        if comp_info:
                            cell = ws.cell(
                                row=row,
                                column=comp_info["col"],
                                value=mark_value,
                            )
                            cell.alignment = Alignment(horizontal="center")
                            cell.border = BORDER

                    # Fill TOT column
                    if col_info.get("tot_col"):
                        tot_cell = ws.cell(
                            row=row, column=col_info["tot_col"], value=total
                        )
                        tot_cell.font = Font(bold=True)
                        tot_cell.alignment = Alignment(horizontal="center")
                        tot_cell.border = BORDER

                # ── Summary columns ──
                total_cell = ws.cell(row=row, column=summary_total_col, value=round(total_marks_sum, 1))
                total_cell.alignment = Alignment(horizontal="center")
                total_cell.border = BORDER

                sgpi_cell = ws.cell(row=row, column=summary_sgpi_col, value=sem_record.sgpa)
                sgpi_cell.font = Font(bold=True, color="2F5496")
                sgpi_cell.alignment = Alignment(horizontal="center")
                sgpi_cell.border = BORDER

                result_str = "PASS" if sem_record.sgpa >= 4.0 else "FAIL"
                result_cell = ws.cell(row=row, column=summary_result_col, value=result_str)
                result_cell.font = Font(
                    bold=True,
                    color="006600" if result_str == "PASS" else "CC0000",
                )
                result_cell.fill = (
                    PatternFill("solid", fgColor="C6EFCE")
                    if result_str == "PASS"
                    else PatternFill("solid", fgColor="FFC7CE")
                )
                result_cell.alignment = Alignment(horizontal="center")
                result_cell.border = BORDER

                # Borders for full row
                for c in range(1, summary_result_col + 1):
                    ws.cell(row=row, column=c).border = BORDER

            # ── Students metadata (hidden row 101) ──
            ws.cell(row=101, column=1, value="__STUDENTS_META__")
            ws.cell(row=101, column=2, value=json.dumps({
                "total_students": len(sorted_students),
                "students_with_marks": students_filled,
                "students_without_marks": students_no_marks,
                "branch": branch,
                "admission_year": adm_year,
                "semester": sem_num,
                "exported_at": datetime.now().isoformat(),
            }))
            ws.row_dimensions[101].hidden = True

            # ── Save file ──
            filename = (
                f"marks_sem{sem_num}_{branch}_{adm_year}_"
                f"{academic_year.replace('-', '_')}.xlsx"
            )
            filepath = os.path.join(output_dir, filename)

            # Handle permission error: close file if open
            try:
                wb.save(filepath)
            except PermissionError:
                # Try alternate filename with timestamp
                ts = datetime.now().strftime("%H%M%S")
                alt_filename = (
                    f"marks_sem{sem_num}_{branch}_{adm_year}_"
                    f"{academic_year.replace('-', '_')}_{ts}.xlsx"
                )
                filepath = os.path.join(output_dir, alt_filename)
                wb.save(filepath)
                print(f"  ⚠️  Original file locked, saved as: {alt_filename}")

            output_files.append(filepath)

            print(
                f"  ✅ {os.path.basename(filepath):55s} | "
                f"{students_filled:3d} with marks, "
                f"{students_no_marks:3d} empty, "
                f"{len(subjects):2d} subjects"
            )

    # ── Generate overview ──
    if output_files:
        _generate_overview(output_dir, groups, group_semesters, students)

    return output_files


def _generate_overview(
    output_dir: str,
    groups: Dict,
    group_semesters: Dict,
    all_students: List,
):
    """Generate overview Excel summarizing all exports"""
    from app.core.curriculum import get_semester_subjects

    THIN = Side(style="thin")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
    HDR_FILL = PatternFill("solid", fgColor="2F5496")

    wb = Workbook()

    # Sheet 1: Summary
    ws = wb.active
    ws.title = "Export Summary"

    headers = [
        "Branch", "Adm. Year", "Semester", "Acad. Year",
        "Students", "With Marks", "Avg SGPA", "Pass %", "Filename",
    ]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.border = BORDER

    row = 2
    for (adm_year, branch), grp in sorted(groups.items()):
        semesters = sorted(group_semesters.get((adm_year, branch), set()))
        for sem_num in semesters:
            subjects = get_semester_subjects(sem_num, adm_year)
            if not subjects:
                continue

            count = 0
            with_marks = 0
            sgpa_sum = 0.0
            pass_count = 0

            for s in grp:
                sr = next(
                    (r for r in s.semester_records if r.semester_number == sem_num),
                    None,
                )
                if sr and sr.subjects:
                    with_marks += 1
                    sgpa_sum += sr.sgpa
                    if sr.sgpa >= 4.0:
                        pass_count += 1
                count += 1

            avg_sgpa = round(sgpa_sum / with_marks, 2) if with_marks > 0 else 0
            pass_pct = round(pass_count / with_marks * 100, 1) if with_marks > 0 else 0

            ay = get_academic_year(sem_num, adm_year)
            filename = f"marks_sem{sem_num}_{branch}_{adm_year}_{ay.replace('-', '_')}.xlsx"

            vals = [branch, adm_year, sem_num, ay, count, with_marks,
                    avg_sgpa, f"{pass_pct}%", filename]
            for ci, v in enumerate(vals, 1):
                ws.cell(row=row, column=ci, value=v).border = BORDER
            row += 1

    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 18

    # Sheet 2: CGPA Distribution
    ws2 = wb.create_sheet("CGPA Distribution")
    dist_headers = ["Branch", "Adm. Year", "Roll Number", "Name",
                    "Current Sem", "CGPA", "Total Credits", "Category"]
    for i, h in enumerate(dist_headers, 1):
        cell = ws2.cell(row=1, column=i, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.border = BORDER

    row = 2
    for s in sorted(all_students, key=lambda x: (x.branch, x.admission_year, x.roll_number)):
        category = _cgpa_category(s.cgpa)
        vals = [
            s.branch, s.admission_year, s.roll_number, s.name,
            s.current_semester, s.cgpa, s.total_credits_earned, category,
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws2.cell(row=row, column=ci, value=v)
            cell.border = BORDER
            if ci == 6:
                if s.cgpa >= 8.0:
                    cell.fill = PatternFill("solid", fgColor="C6EFCE")
                elif s.cgpa >= 6.0:
                    cell.fill = PatternFill("solid", fgColor="FFEB9C")
                elif s.cgpa > 0:
                    cell.fill = PatternFill("solid", fgColor="FFC7CE")
        row += 1

    for ci in range(1, len(dist_headers) + 1):
        ws2.column_dimensions[get_column_letter(ci)].width = 18

    overview_path = os.path.join(output_dir, "00_OVERVIEW.xlsx")
    try:
        wb.save(overview_path)
        print(f"\n  📋 Overview: {overview_path}")
    except PermissionError:
        ts = datetime.now().strftime("%H%M%S")
        alt = os.path.join(output_dir, f"00_OVERVIEW_{ts}.xlsx")
        wb.save(alt)
        print(f"\n  📋 Overview: {alt} (original locked)")


def _cgpa_category(cgpa: float) -> str:
    if cgpa >= 9.0: return "Outstanding"
    if cgpa >= 8.0: return "Excellent"
    if cgpa >= 7.0: return "Very Good"
    if cgpa >= 6.0: return "Good"
    if cgpa >= 5.0: return "Average"
    if cgpa >= 4.0: return "Below Average"
    if cgpa > 0:    return "At Risk"
    return "No Data"


# ══════════════════════════════════════════════════════════
# 7.  CLI
# ══════════════════════════════════════════════════════════

def print_stats(stats: Dict[str, Any]):
    print(f"\n{'='*70}")
    print(f"📊 SEEDING STATISTICS")
    print(f"{'='*70}")
    print(f"  Total Students:          {stats['total_students']}")
    print(f"  Updated:                 {stats['updated']}")
    print(f"  Skipped (already had):   {stats['skipped']}")
    print(f"  Semesters Created:       {stats['semesters_created']}")
    print(f"  Semesters Skipped:       {stats['semesters_skipped_existing']}")

    if stats["profile_distribution"]:
        print(f"\n  📈 Performance Profile Distribution:")
        for name, count in sorted(stats["profile_distribution"].items(), key=lambda x: -x[1]):
            bar = "█" * max(1, count // 2)
            print(f"    {name:18s}: {count:4d}  {bar}")

    total_grades = sum(stats["grade_distribution"].values())
    if total_grades > 0:
        print(f"\n  📊 Grade Distribution ({total_grades} total):")
        for grade in ["O", "A+", "A", "B+", "B", "C", "P", "F"]:
            count = stats["grade_distribution"].get(grade, 0)
            pct = (count / total_grades * 100)
            bar = "█" * max(1, count // 10)
            print(f"    {grade:4s}: {count:6d}  ({pct:5.1f}%)  {bar}")
    else:
        print(f"\n  ℹ️  No new grades generated (all semesters already existed)")
        print(f"      Use --overwrite to regenerate marks")

    if stats["by_admission_year"]:
        print(f"\n  📅 By Admission Year:")
        for yr, data in sorted(stats["by_admission_year"].items()):
            print(
                f"    {yr}: {data['students']:3d} students, "
                f"{data['semesters_total']:3d} new semesters, "
                f"avg CGPA: {data['avg_cgpa']:.2f}"
            )
    print(f"{'='*70}\n")


async def main():
    parser = argparse.ArgumentParser(
        description=(
            "Generate comprehensive marks for all students and export "
            "to university-format XLSX (compatible with bulk upload)"
        )
    )
    parser.add_argument("--branch", "-b", type=str, default=None,
                        help="Filter by branch (e.g., IT, COMP)")
    parser.add_argument("--admission-year", "-a", type=int, default=None,
                        help="Filter by admission year")
    parser.add_argument("--seed-only", action="store_true",
                        help="Only seed marks, don't export")
    parser.add_argument("--export-only", action="store_true",
                        help="Only export existing marks, don't seed")
    parser.add_argument("--overwrite", action="store_true",
                        help="Overwrite existing semester records with fresh marks")
    parser.add_argument("--dry-run", action="store_true",
                        help="Generate marks but don't save to DB")
    parser.add_argument("--output-dir", "-o", type=str,
                        default="./exported_marks",
                        help="Output directory for XLSX files")

    args = parser.parse_args()
    client = await init_database()

    try:
        # ── SEED ──
        if not args.export_only:
            print(f"\n{'='*70}")
            print(f"🎲 SEEDING MARKS FOR ALL STUDENTS")
            if args.branch:
                print(f"   Branch: {args.branch}")
            if args.admission_year:
                print(f"   Admission Year: {args.admission_year}")
            if args.overwrite:
                print(f"   ⚠️  OVERWRITE mode: will regenerate ALL marks")
            else:
                print(f"   ℹ️  SAFE mode: only fills empty semesters")
            if args.dry_run:
                print(f"   🧪 DRY RUN — no DB writes")
            print(f"{'='*70}\n")

            stats = await seed_marks_for_all_students(
                branch_filter=args.branch,
                admission_year_filter=args.admission_year,
                overwrite=args.overwrite,
                dry_run=args.dry_run,
            )
            print_stats(stats)

        # ── EXPORT ──
        if not args.seed_only:
            print(f"\n{'='*70}")
            print(f"📤 EXPORTING MARKS (University Template Format)")
            print(f"   Output: {os.path.abspath(args.output_dir)}")
            print(f"{'='*70}\n")

            files = await export_all_marks_to_xlsx(
                branch_filter=args.branch,
                admission_year_filter=args.admission_year,
                output_dir=args.output_dir,
            )

            if files:
                print(f"\n✅ Exported {len(files)} file(s) to: {os.path.abspath(args.output_dir)}")
                print(f"\n💡 These files can be re-uploaded via Admin > Bulk Marks Upload")
                print(f"   Each file has the __COLUMN_MAP__ embedded for proper parsing")

    finally:
        client.close()
        print("\n🔌 Database connection closed")


if __name__ == "__main__":
    asyncio.run(main())