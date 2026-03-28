# academic-advisor-backend/scripts/seed_marks_data.py
"""
Seed random marks data for all students
Run: python -m scripts.seed_marks_data
"""

import asyncio
import sys
import os
import random
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


async def init_database():
    """Initialize database connection"""
    from motor.motor_asyncio import AsyncIOMotorClient
    from beanie import init_beanie
    from app.core.config import settings
    from app.models.student_profile import StudentProfile, SemesterRecord, SubjectScore
    from app.models.pending_marks import PendingStudentMarks
    
    client = AsyncIOMotorClient(settings.MONGODB_URL)
    db = client[settings.DATABASE_NAME]
    
    await init_beanie(
        database=db,
        document_models=[StudentProfile, PendingStudentMarks]
    )
    
    print("✅ Database connected")
    return client


def generate_random_marks(internal_max: float, external_max: float, pass_rate: float = 0.85):
    """
    Generate random marks with realistic distribution.
    
    pass_rate: probability of passing (0.0 to 1.0)
    """
    # Decide if this student passes or fails this subject
    is_passing = random.random() < pass_rate
    
    if is_passing:
        # Passing marks: 40% to 100% of max
        # With bell curve distribution (most scores around 60-75%)
        internal_pct = min(100, max(40, random.gauss(65, 15))) / 100
        external_pct = min(100, max(40, random.gauss(62, 18))) / 100
    else:
        # Failing marks: 10% to 39% of max
        internal_pct = random.uniform(0.15, 0.38)
        external_pct = random.uniform(0.10, 0.38)
    
    internal = round(internal_max * internal_pct, 1)
    external = round(external_max * external_pct, 1)
    
    return internal, external


def calculate_grade(total: float, max_marks: float) -> dict:
    """Calculate grade and grade points"""
    if max_marks <= 0:
        max_marks = 100
    pct = (total / max_marks) * 100
    
    if pct >= 90: return {"grade": "O", "points": 10.0}
    if pct >= 80: return {"grade": "A+", "points": 9.0}
    if pct >= 70: return {"grade": "A", "points": 8.0}
    if pct >= 60: return {"grade": "B+", "points": 7.0}
    if pct >= 50: return {"grade": "B", "points": 6.0}
    if pct >= 45: return {"grade": "C", "points": 5.0}
    if pct >= 40: return {"grade": "P", "points": 4.0}
    return {"grade": "F", "points": 0.0}


def get_academic_year(semester: int, admission_year: int) -> str:
    """Calculate academic year for a semester"""
    year_offset = (semester - 1) // 2
    start_year = admission_year + year_offset
    return f"{start_year}-{str(start_year + 1)[2:]}"


async def create_sample_students(branch: str, admission_year: int, count: int = 10):
    """Create sample students if not enough exist"""
    from app.models.student_profile import StudentProfile
    
    existing = await StudentProfile.find({
        "branch": {"$regex": f"^{branch}$", "$options": "i"},
        "admission_year": admission_year,
    }).count()
    
    if existing >= count:
        print(f"✅ Already have {existing} students for {branch} {admission_year}")
        return
    
    to_create = count - existing
    print(f"📝 Creating {to_create} sample students for {branch} {admission_year}...")
    
    # Sample Indian names
    first_names = [
        "Aarav", "Vivaan", "Aditya", "Vihaan", "Arjun", "Reyansh", "Ayaan", "Krishna",
        "Ishaan", "Shaurya", "Atharv", "Advait", "Dhruv", "Kabir", "Ritvik", "Aarush",
        "Aanya", "Saanvi", "Ananya", "Pari", "Myra", "Sara", "Aadhya", "Avni",
        "Kiara", "Diya", "Prisha", "Anvi", "Ira", "Navya", "Anika", "Riya",
        "Rohan", "Rahul", "Raj", "Amit", "Vikram", "Sanjay", "Priya", "Neha",
        "Pooja", "Sneha", "Anjali", "Kavya", "Tanvi", "Shruti", "Nikita", "Megha"
    ]
    
    last_names = [
        "Sharma", "Patel", "Singh", "Kumar", "Gupta", "Shah", "Joshi", "Verma",
        "Mehta", "Reddy", "Nair", "Iyer", "Rao", "Pillai", "Menon", "Desai",
        "Patil", "Kulkarni", "Jain", "Agarwal", "Mishra", "Pandey", "Tiwari", "Yadav",
        "Chauhan", "Malhotra", "Kapoor", "Khanna", "Bhatia", "Sinha", "Das", "Roy"
    ]
    
    # Calculate current semester
    current_date = datetime.now()
    current_year = current_date.year
    current_month = current_date.month
    
    if current_month >= 7:
        year_diff = current_year - admission_year
        current_sem = (year_diff * 2) + 1
    else:
        year_diff = current_year - admission_year
        current_sem = year_diff * 2
    
    current_sem = max(1, min(current_sem, 8))
    academic_year = get_academic_year(current_sem, admission_year)
    
    created = 0
    for i in range(to_create):
        roll_num = existing + i + 1
        roll_number = f"{admission_year}{branch}{roll_num:03d}"
        
        first = random.choice(first_names)
        last = random.choice(last_names)
        name = f"{first} {last}"
        
        # Generate random 5-digit seat number
        seat_number = str(random.randint(10000, 99999))
        
        profile = StudentProfile(
            user_id=f"pending_{roll_number}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{i}",
            name=name,
            roll_number=roll_number,
            current_seat_number=seat_number,
            branch=branch,
            admission_year=admission_year,
            email=f"{first.lower()}.{last.lower()}{roll_num}@college.edu",
            current_semester=current_sem,
            current_academic_year=academic_year,
            cgpa=0.0,
            total_credits_earned=0,
            total_credits_required=160,
            semester_records=[],
            skills=[],
            interests=[],
            career_goals=[],
            created_at=datetime.now(),
            last_updated=datetime.now(),
        )
        
        await profile.insert()
        created += 1
    
    print(f"✅ Created {created} new students")


async def generate_marks_for_student(student, semesters: list, pass_rate: float = 0.85):
    """Generate random marks for a student for specified semesters"""
    from app.models.student_profile import SemesterRecord, SubjectScore
    from app.core.curriculum import get_semester_subjects
    
    updated = False
    
    for sem_num in semesters:
        # Check if semester already has marks
        existing_sem = next(
            (sr for sr in student.semester_records if sr.semester_number == sem_num),
            None
        )
        
        if existing_sem and len(existing_sem.subjects) > 0:
            # Already has marks, skip
            continue
        
        # Get subjects for this semester
        subjects = get_semester_subjects(sem_num, student.admission_year)
        
        if not subjects:
            print(f"  ⚠️ No subjects for semester {sem_num}, admission {student.admission_year}")
            continue
        
        academic_year = get_academic_year(sem_num, student.admission_year)
        
        # Generate marks for each subject
        subject_scores = []
        total_grade_points = 0.0
        total_credits = 0
        credits_earned = 0
        
        for sub in subjects:
            # Generate random marks
            internal, external = generate_random_marks(
                sub.internal_max, 
                sub.external_max,
                pass_rate
            )
            
            total = internal + external
            max_total = sub.internal_max + sub.external_max
            grade_info = calculate_grade(total, max_total)
            
            score = SubjectScore(
                subject_code=sub.subject_code,
                subject_name=sub.subject_name,
                credits=sub.credits,
                internal_marks=internal,
                external_marks=external,
                total_marks=total,
                grade=grade_info["grade"],
                grade_points=grade_info["points"],
                is_elective=sub.is_elective,
                is_practical=sub.is_practical,
            )
            
            subject_scores.append(score)
            
            total_credits += sub.credits
            if grade_info["grade"] != "F":
                total_grade_points += grade_info["points"] * sub.credits
                credits_earned += sub.credits
        
        # Calculate SGPA
        sgpa = round(total_grade_points / total_credits, 2) if total_credits > 0 else 0.0
        
        # Create semester record
        sem_record = SemesterRecord(
            semester_number=sem_num,
            academic_year=academic_year,
            subjects=subject_scores,
            sgpa=sgpa,
            total_credits=total_credits,
            credits_earned=credits_earned,
            is_complete=True,
            created_at=datetime.now(),
        )
        
        # Add or update semester record
        if existing_sem:
            idx = student.semester_records.index(existing_sem)
            student.semester_records[idx] = sem_record
        else:
            student.semester_records.append(sem_record)
        
        updated = True
        print(f"    📝 Sem {sem_num}: SGPA={sgpa:.2f}, Credits={credits_earned}/{total_credits}")
    
    if updated:
        # Sort semester records
        student.semester_records.sort(key=lambda x: x.semester_number)
        
        # Recalculate CGPA
        all_gp = sum(sr.sgpa * sr.total_credits for sr in student.semester_records if sr.is_complete)
        all_creds = sum(sr.total_credits for sr in student.semester_records if sr.is_complete)
        all_earned = sum(sr.credits_earned for sr in student.semester_records)
        
        student.cgpa = round(all_gp / all_creds, 2) if all_creds > 0 else 0.0
        student.total_credits_earned = all_earned
        student.marks_synced_at = datetime.now()
        student.last_updated = datetime.now()
        
        await student.save()
        
        return True
    
    return False


async def seed_all_marks():
    """Seed random marks for all students"""
    from app.models.student_profile import StudentProfile
    
    client = await init_database()
    
    try:
        # ═══════════════════════════════════════════════════
        # CONFIGURATION
        # ═══════════════════════════════════════════════════
        
        BRANCH = "IT"
        ADMISSION_YEAR = 2022
        SEMESTERS_TO_SEED = [1, 2, 3, 4, 5]  # Which semesters to generate marks for
        MIN_STUDENTS = 20  # Minimum number of students to have
        PASS_RATE = 0.85  # 85% pass rate
        
        # ═══════════════════════════════════════════════════
        
        print(f"\n{'='*70}")
        print(f"🎲 SEEDING RANDOM MARKS DATA")
        print(f"{'='*70}")
        print(f"Branch: {BRANCH}")
        print(f"Admission Year: {ADMISSION_YEAR}")
        print(f"Semesters: {SEMESTERS_TO_SEED}")
        print(f"Pass Rate: {PASS_RATE*100:.0f}%")
        print(f"{'='*70}\n")
        
        # Step 1: Create sample students if needed
        await create_sample_students(BRANCH, ADMISSION_YEAR, MIN_STUDENTS)
        
        # Step 2: Fetch all students
        students = await StudentProfile.find({
            "branch": {"$regex": f"^{BRANCH}$", "$options": "i"},
            "admission_year": ADMISSION_YEAR,
        }).sort("roll_number").to_list()
        
        print(f"\n👥 Found {len(students)} students")
        print(f"{'='*70}\n")
        
        # Step 3: Generate marks for each student
        updated_count = 0
        
        for i, student in enumerate(students, 1):
            print(f"[{i}/{len(students)}] {student.roll_number} - {student.name}")
            
            result = await generate_marks_for_student(
                student, 
                SEMESTERS_TO_SEED, 
                PASS_RATE
            )
            
            if result:
                updated_count += 1
                print(f"    ✅ CGPA: {student.cgpa:.2f}")
            else:
                print(f"    ⏭️ Already has marks")
        
        print(f"\n{'='*70}")
        print(f"✅ SEEDING COMPLETE!")
        print(f"{'='*70}")
        print(f"Total Students: {len(students)}")
        print(f"Updated: {updated_count}")
        print(f"Semesters Seeded: {SEMESTERS_TO_SEED}")
        print(f"{'='*70}\n")
        
    finally:
        client.close()


async def seed_marks_for_multiple_batches():
    """Seed marks for multiple batches/branches"""
    from app.models.student_profile import StudentProfile
    
    client = await init_database()
    
    try:
        # ═══════════════════════════════════════════════════
        # CONFIGURATION - Multiple batches
        # ═══════════════════════════════════════════════════
        
        BATCHES = [
            {"branch": "IT", "admission_year": 2022, "semesters": [1, 2, 3, 4, 5], "count": 25},
            {"branch": "IT", "admission_year": 2023, "semesters": [1, 2, 3], "count": 30},
            {"branch": "IT", "admission_year": 2024, "semesters": [1], "count": 35},
            {"branch": "COMP", "admission_year": 2022, "semesters": [1, 2, 3, 4, 5], "count": 20},
        ]
        
        PASS_RATE = 0.85
        
        # ═══════════════════════════════════════════════════
        
        print(f"\n{'='*70}")
        print(f"🎲 SEEDING MARKS FOR MULTIPLE BATCHES")
        print(f"{'='*70}\n")
        
        total_updated = 0
        
        for batch in BATCHES:
            branch = batch["branch"]
            adm_year = batch["admission_year"]
            semesters = batch["semesters"]
            count = batch["count"]
            
            print(f"\n📚 Processing {branch} - {adm_year}")
            print(f"   Semesters: {semesters}")
            print(f"-" * 50)
            
            # Create students if needed
            await create_sample_students(branch, adm_year, count)
            
            # Fetch students
            students = await StudentProfile.find({
                "branch": {"$regex": f"^{branch}$", "$options": "i"},
                "admission_year": adm_year,
            }).sort("roll_number").to_list()
            
            # Generate marks
            batch_updated = 0
            for student in students:
                result = await generate_marks_for_student(student, semesters, PASS_RATE)
                if result:
                    batch_updated += 1
            
            print(f"   ✅ Updated {batch_updated}/{len(students)} students")
            total_updated += batch_updated
        
        print(f"\n{'='*70}")
        print(f"✅ ALL BATCHES COMPLETE!")
        print(f"Total Updated: {total_updated}")
        print(f"{'='*70}\n")
        
    finally:
        client.close()


async def export_marks_to_excel():
    """Export all marks to Excel after seeding"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from app.models.student_profile import StudentProfile
    from app.core.curriculum import get_semester_subjects
    
    client = await init_database()
    
    try:
        BRANCH = "IT"
        ADMISSION_YEAR = 2022
        OUTPUT_DIR = "./exported_marks"
        
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        
        # Fetch students
        students = await StudentProfile.find({
            "branch": {"$regex": f"^{BRANCH}$", "$options": "i"},
            "admission_year": ADMISSION_YEAR,
        }).sort("roll_number").to_list()
        
        print(f"\n📊 Exporting marks for {len(students)} students")
        
        # Find all semesters with marks
        semesters_with_marks = set()
        for student in students:
            for sr in student.semester_records:
                semesters_with_marks.add(sr.semester_number)
        
        print(f"Semesters with marks: {sorted(semesters_with_marks)}")
        
        # Export each semester
        for sem in sorted(semesters_with_marks):
            wb = Workbook()
            ws = wb.active
            ws.title = f"Semester {sem}"
            
            # Styles
            THIN = Side(style="thin")
            BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            hdr_font = Font(bold=True, color="FFFFFF")
            hdr_fill = PatternFill("solid", fgColor="2F5496")
            pass_fill = PatternFill("solid", fgColor="C6EFCE")
            fail_fill = PatternFill("solid", fgColor="FFC7CE")
            
            # Get subjects for this semester
            subjects = get_semester_subjects(sem, ADMISSION_YEAR)
            
            # Headers
            headers = ["Sr.No", "Roll Number", "Name", "Seat No"]
            for sub in subjects:
                headers.extend([f"{sub.subject_code}_INT", f"{sub.subject_code}_EXT", f"{sub.subject_code}_TOT", f"{sub.subject_code}_GR"])
            headers.extend(["Total Marks", "SGPA", "Credits Earned", "Result"])
            
            for i, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=i, value=h)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.border = BORDER
                cell.alignment = Alignment(horizontal="center")
            
            # Data rows
            row = 2
            for idx, student in enumerate(students, 1):
                # Find semester record
                sem_record = next((sr for sr in student.semester_records if sr.semester_number == sem), None)
                
                if not sem_record:
                    continue
                
                col = 1
                ws.cell(row=row, column=col, value=idx).border = BORDER
                col += 1
                
                ws.cell(row=row, column=col, value=student.roll_number).border = BORDER
                col += 1
                
                ws.cell(row=row, column=col, value=student.name).border = BORDER
                col += 1
                
                ws.cell(row=row, column=col, value=student.current_seat_number or "").border = BORDER
                col += 1
                
                total_marks = 0
                
                # Subject marks
                for sub in subjects:
                    # Find matching subject in semester record
                    subj_score = next(
                        (s for s in sem_record.subjects if s.subject_code == sub.subject_code),
                        None
                    )
                    
                    if subj_score:
                        ws.cell(row=row, column=col, value=subj_score.internal_marks).border = BORDER
                        col += 1
                        ws.cell(row=row, column=col, value=subj_score.external_marks).border = BORDER
                        col += 1
                        ws.cell(row=row, column=col, value=subj_score.total_marks).border = BORDER
                        col += 1
                        
                        gr_cell = ws.cell(row=row, column=col, value=subj_score.grade)
                        gr_cell.border = BORDER
                        gr_cell.font = Font(bold=True)
                        if subj_score.grade == "F":
                            gr_cell.fill = fail_fill
                        else:
                            gr_cell.fill = pass_fill
                        col += 1
                        
                        total_marks += subj_score.total_marks
                    else:
                        # Empty cells
                        for _ in range(4):
                            ws.cell(row=row, column=col, value="").border = BORDER
                            col += 1
                
                # Summary
                ws.cell(row=row, column=col, value=total_marks).border = BORDER
                col += 1
                
                sgpa_cell = ws.cell(row=row, column=col, value=sem_record.sgpa)
                sgpa_cell.border = BORDER
                sgpa_cell.font = Font(bold=True)
                col += 1
                
                ws.cell(row=row, column=col, value=sem_record.credits_earned).border = BORDER
                col += 1
                
                result = "PASS" if sem_record.sgpa >= 4.0 else "FAIL"
                result_cell = ws.cell(row=row, column=col, value=result)
                result_cell.border = BORDER
                result_cell.font = Font(bold=True, color="006600" if result == "PASS" else "CC0000")
                result_cell.fill = pass_fill if result == "PASS" else fail_fill
                
                row += 1
            
            # Adjust column widths
            for i in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(i)].width = 12
            ws.column_dimensions["B"].width = 18
            ws.column_dimensions["C"].width = 25
            
            # Save
            filename = f"marks_sem{sem}_{BRANCH}_{ADMISSION_YEAR}.xlsx"
            filepath = os.path.join(OUTPUT_DIR, filename)
            wb.save(filepath)
            print(f"  ✅ Exported: {filepath}")
        
        print(f"\n✅ Export complete! Files saved to: {os.path.abspath(OUTPUT_DIR)}")
        
    finally:
        client.close()


async def show_summary():
    """Show summary of all marks in database"""
    from app.models.student_profile import StudentProfile
    
    client = await init_database()
    
    try:
        # Get all students
        students = await StudentProfile.find({}).to_list()
        
        print(f"\n{'='*70}")
        print(f"📊 MARKS DATABASE SUMMARY")
        print(f"{'='*70}\n")
        
        # Group by branch and admission year
        groups = {}
        for s in students:
            key = (s.branch, s.admission_year)
            if key not in groups:
                groups[key] = {"students": 0, "with_marks": 0, "semesters": set()}
            groups[key]["students"] += 1
            if s.semester_records:
                groups[key]["with_marks"] += 1
                for sr in s.semester_records:
                    groups[key]["semesters"].add(sr.semester_number)
        
        print(f"{'Branch':<10} {'Admission':<12} {'Students':<10} {'With Marks':<12} {'Semesters'}")
        print("-" * 70)
        
        for (branch, adm_year), data in sorted(groups.items()):
            sems = ", ".join(str(s) for s in sorted(data["semesters"])) if data["semesters"] else "None"
            print(f"{branch:<10} {adm_year:<12} {data['students']:<10} {data['with_marks']:<12} {sems}")
        
        print(f"\nTotal Students: {len(students)}")
        print(f"With Marks: {sum(1 for s in students if s.semester_records)}")
        print()
        
    finally:
        client.close()


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Seed random marks data")
    parser.add_argument("--seed", action="store_true", help="Seed marks for one batch (IT 2022)")
    parser.add_argument("--seed-all", action="store_true", help="Seed marks for multiple batches")
    parser.add_argument("--export", action="store_true", help="Export marks to Excel")
    parser.add_argument("--summary", action="store_true", help="Show database summary")
    parser.add_argument("--full", action="store_true", help="Seed all batches + Export (complete workflow)")
    
    args = parser.parse_args()
    
    if args.seed:
        asyncio.run(seed_all_marks())
    elif args.seed_all:
        asyncio.run(seed_marks_for_multiple_batches())
    elif args.export:
        asyncio.run(export_marks_to_excel())
    elif args.summary:
        asyncio.run(show_summary())
    elif args.full:
        print("🚀 Running full workflow: Seed + Export")
        asyncio.run(seed_marks_for_multiple_batches())
        asyncio.run(export_marks_to_excel())
    else:
        # Default: show help
        parser.print_help()
        print("\n📌 Quick start:")
        print("  python -m scripts.seed_marks_data --full")