# academic-advisor-backend/scripts/export_marks_template.py
"""
Export marks template with existing data
Run: python -m scripts.export_marks_template
"""

import asyncio
import sys
import os

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


async def init_database():
    """Initialize database connection properly"""
    from motor.motor_asyncio import AsyncIOMotorClient
    from beanie import init_beanie
    from app.core.config import settings
    from app.models.student_profile import StudentProfile
    from app.models.pending_marks import PendingStudentMarks
    
    # Create motor client
    client = AsyncIOMotorClient(settings.MONGODB_URL)
    db = client[settings.DATABASE_NAME]
    
    # Initialize beanie with document models
    await init_beanie(
        database=db,
        document_models=[
            StudentProfile,
            PendingStudentMarks,
        ]
    )
    
    print("✅ Database connected")
    return client


async def export_marks():
    """Export marks for a single semester"""
    
    # Initialize database
    client = await init_database()
    
    try:
        from app.services.bulk_marks_service import bulk_marks_service
        
        # ═══════════════════════════════════════════════════
        # CONFIGURATION - CHANGE THESE VALUES
        # ═══════════════════════════════════════════════════
        
        SEMESTER = 5
        BRANCH = "IT"
        ADMISSION_YEAR = 2022
        ACADEMIC_YEAR = "2024-25"
        OUTPUT_DIR = "./exported_marks"
        
        # ═══════════════════════════════════════════════════
        
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        
        print(f"\n{'='*60}")
        print(f"📊 Exporting Marks for Semester {SEMESTER}")
        print(f"{'='*60}")
        print(f"Branch: {BRANCH}")
        print(f"Admission Year: {ADMISSION_YEAR}")
        print(f"Academic Year: {ACADEMIC_YEAR}")
        print(f"{'='*60}\n")
        
        # Check if the method exists, if not use the alternative approach
        if hasattr(bulk_marks_service, 'generate_template_with_marks'):
            buf = await bulk_marks_service.generate_template_with_marks(
                semester=SEMESTER,
                branch=BRANCH,
                academic_year=ACADEMIC_YEAR,
                admission_year=ADMISSION_YEAR,
            )
        else:
            # Use the direct generation method
            buf = await generate_marks_template_direct(
                semester=SEMESTER,
                branch=BRANCH,
                academic_year=ACADEMIC_YEAR,
                admission_year=ADMISSION_YEAR,
            )
        
        filename = f"marks_export_sem{SEMESTER}_{BRANCH}_{ACADEMIC_YEAR.replace('-', '_')}.xlsx"
        filepath = os.path.join(OUTPUT_DIR, filename)
        
        with open(filepath, 'wb') as f:
            f.write(buf.read())
        
        print(f"✅ Exported: {filepath}")
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
    finally:
        client.close()
    
    print(f"\n{'='*60}")
    print(f"✅ Export complete!")
    print(f"{'='*60}\n")


async def generate_marks_template_direct(
    semester: int,
    branch: str,
    academic_year: str,
    admission_year: int,
):
    """
    Directly generate Excel template with marks data.
    This is a standalone implementation.
    """
    import io
    import json
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    from app.models.student_profile import StudentProfile
    from app.core.curriculum import get_semester_subjects, get_elective_options
    
    # Styles
    THIN = Side(style="thin")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    
    # Get subjects for this semester
    subjects = get_semester_subjects(semester, admission_year)
    if not subjects:
        raise ValueError(f"No subjects found for semester {semester}, admission year {admission_year}")
    
    print(f"📚 Found {len(subjects)} subjects for semester {semester}")
    
    # Fetch students
    query_filter = {
        "branch": {"$regex": f"^{branch}$", "$options": "i"},
        "admission_year": admission_year,
    }
    
    students = await StudentProfile.find(query_filter).sort("roll_number").to_list()
    print(f"👥 Found {len(students)} students")
    
    # Build workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Marks Data"
    
    # ── Row 1: Metadata ──
    meta_font = Font(bold=True, size=11)
    meta_fill = PatternFill("solid", fgColor="D6E4F0")
    
    meta = [
        ("semester", semester), ("branch", branch),
        ("academic_year", academic_year), ("admission_year", admission_year),
    ]
    col = 1
    for key, val in meta:
        ws.cell(row=1, column=col, value=key).font = meta_font
        ws.cell(row=1, column=col).fill = meta_fill
        ws.cell(row=1, column=col + 1, value=val).font = meta_font
        ws.cell(row=1, column=col + 1).fill = meta_fill
        col += 2
    
    # ── Row 2: Instructions ──
    ws.cell(row=2, column=1, value="This template contains existing marks data. Edit and re-upload to update.")
    ws.cell(row=2, column=1).font = Font(italic=True, color="555555")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=15)
    
    # ── Row 4: Headers ──
    HDR_ROW = 4
    COMP_ROW = 5
    MAX_ROW = 6
    DATA_START = 7
    
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    comp_font = Font(bold=True, size=10)
    comp_fill = PatternFill("solid", fgColor="B4C6E7")
    max_font = Font(italic=True, size=9, color="666666")
    max_fill = PatternFill("solid", fgColor="F2F2F2")
    
    # Fixed columns
    fixed_headers = [("Sr.No", 6), ("Roll Number", 18), ("Name", 28)]
    for i, (label, width) in enumerate(fixed_headers, 1):
        cell = ws.cell(row=HDR_ROW, column=i, value=label)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        ws.merge_cells(start_row=HDR_ROW, start_column=i, end_row=COMP_ROW, end_column=i)
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Subject columns
    col = len(fixed_headers) + 1
    subject_col_map = {}  # Maps subject_code -> {component -> column}
    
    # Category colors
    cat_fills = {
        "PCC": PatternFill("solid", fgColor="2F5496"),
        "BSC": PatternFill("solid", fgColor="375623"),
        "ESC": PatternFill("solid", fgColor="548235"),
        "AEC": PatternFill("solid", fgColor="BF8F00"),
        "PEC": PatternFill("solid", fgColor="7030A0"),
        "OEC": PatternFill("solid", fgColor="00B0F0"),
        "LBC": PatternFill("solid", fgColor="C55A11"),
        "SBL": PatternFill("solid", fgColor="ED7D31"),
        "MNP": PatternFill("solid", fgColor="FF0066"),
        "MJP": PatternFill("solid", fgColor="CC0000"),
        "INT": PatternFill("solid", fgColor="002060"),
    }
    
    for sub in subjects:
        subject_col_map[sub.subject_code] = {"start": col, "components": {}}
        
        # Determine components based on subject type
        ct = sub.course_type
        if ct in ("MNP", "MJP", "INT"):
            components = [("TW", sub.internal_max)]
        elif ct == "SBL":
            components = [("TW", sub.internal_max), ("PR", sub.external_max)]
        elif ct == "LBC" or sub.is_practical:
            components = [("IA", sub.internal_max), ("PR", sub.external_max)]
        else:
            # Theory: CA, MSE, ESE
            ca = sub.internal_max
            if sub.external_max == 80:
                mse, ese = 30, 50
            elif sub.external_max == 0:
                components = [("CA", ca)]
                mse, ese = 0, 0
            else:
                mse = round(sub.external_max * 0.375)
                ese = sub.external_max - mse
            if sub.external_max > 0:
                components = [("CA", ca), ("MSE", mse), ("ESE", ese)]
        
        n_cols = len(components) + 1  # +1 for TOT
        start_col = col
        
        # Subject header (merged)
        header_text = f"{sub.subject_name}\n({sub.subject_code}) [{sub.credits} cr]"
        cell = ws.cell(row=HDR_ROW, column=start_col, value=header_text)
        cell.font = hdr_font
        cell.fill = cat_fills.get(sub.course_type, hdr_fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        
        if n_cols > 1:
            ws.merge_cells(start_row=HDR_ROW, start_column=start_col, end_row=HDR_ROW, end_column=start_col + n_cols - 1)
        
        # Component headers
        ci = start_col
        for comp_name, comp_max in components:
            # Component name
            cc = ws.cell(row=COMP_ROW, column=ci, value=comp_name)
            cc.font = comp_font
            cc.fill = comp_fill
            cc.alignment = Alignment(horizontal="center")
            cc.border = BORDER
            
            # Max marks
            mc = ws.cell(row=MAX_ROW, column=ci, value=comp_max)
            mc.font = max_font
            mc.fill = max_fill
            mc.alignment = Alignment(horizontal="center")
            mc.border = BORDER
            
            ws.column_dimensions[get_column_letter(ci)].width = 8
            subject_col_map[sub.subject_code]["components"][comp_name] = ci
            ci += 1
        
        # TOT column
        tc = ws.cell(row=COMP_ROW, column=ci, value="TOT")
        tc.font = Font(bold=True, size=10)
        tc.fill = comp_fill
        tc.alignment = Alignment(horizontal="center")
        tc.border = BORDER
        
        total_max = sum(c[1] for c in components)
        tm = ws.cell(row=MAX_ROW, column=ci, value=total_max)
        tm.font = max_font
        tm.fill = max_fill
        tm.alignment = Alignment(horizontal="center")
        tm.border = BORDER
        
        ws.column_dimensions[get_column_letter(ci)].width = 8
        subject_col_map[sub.subject_code]["tot_col"] = ci
        
        col = ci + 1
    
    # Summary columns
    summary_start = col
    for label in ["Total", "SGPA", "Result"]:
        cell = ws.cell(row=HDR_ROW, column=col, value=label)
        cell.font = hdr_font
        cell.fill = PatternFill("solid", fgColor="DDEBF7")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        ws.merge_cells(start_row=HDR_ROW, start_column=col, end_row=COMP_ROW, end_column=col)
        ws.column_dimensions[get_column_letter(col)].width = 10
        col += 1
    
    total_cols = col - 1
    
    # Max row for fixed columns
    for i in range(1, len(fixed_headers) + 1):
        ws.cell(row=MAX_ROW, column=i, value="").fill = max_fill
        ws.cell(row=MAX_ROW, column=i).border = BORDER
    
    # ── Fill student data with marks ──
    students_with_marks = 0
    students_without_marks = 0
    
    for idx, student in enumerate(students):
        row = DATA_START + idx
        
        # Sr. No
        ws.cell(row=row, column=1, value=idx + 1).alignment = Alignment(horizontal="center")
        
        # Roll Number
        roll_cell = ws.cell(row=row, column=2, value=student.roll_number)
        roll_cell.font = Font(bold=True)
        roll_cell.alignment = Alignment(horizontal="center")
        
        # Name
        ws.cell(row=row, column=3, value=student.name)
        
        # Find semester record
        sem_record = None
        for sr in student.semester_records:
            if sr.semester_number == semester:
                sem_record = sr
                break
        
        if sem_record:
            students_with_marks += 1
            
            # Highlight row with marks (light green)
            for c in range(1, 4):
                ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor="E2EFDA")
            
            total_marks = 0
            
            # Fill marks for each subject
            for subj in sem_record.subjects:
                sub_code = subj.subject_code
                
                # Find matching subject in curriculum (might have different code for electives)
                col_info = subject_col_map.get(sub_code)
                
                if not col_info:
                    # Try to find by elective group
                    for curr_sub in subjects:
                        if curr_sub.is_elective and curr_sub.elective_group:
                            # Check if this subject's code matches any elective option
                            options = get_elective_options(curr_sub.elective_group)
                            for opt in options:
                                if opt["code"].upper() == sub_code.upper():
                                    col_info = subject_col_map.get(curr_sub.subject_code)
                                    break
                        if col_info:
                            break
                
                if not col_info:
                    print(f"  ⚠️ Subject {sub_code} not found in curriculum for {student.roll_number}")
                    continue
                
                internal = subj.internal_marks
                external = subj.external_marks
                total = subj.total_marks
                total_marks += total
                
                # Fill component marks based on subject type
                comps = col_info["components"]
                
                if "TW" in comps and "PR" not in comps:
                    # Project/Internship
                    ws.cell(row=row, column=comps["TW"], value=internal)
                elif "TW" in comps and "PR" in comps:
                    # Skill Lab
                    ws.cell(row=row, column=comps["TW"], value=internal)
                    ws.cell(row=row, column=comps["PR"], value=external)
                elif "IA" in comps:
                    # Lab
                    ws.cell(row=row, column=comps["IA"], value=internal)
                    if "PR" in comps:
                        ws.cell(row=row, column=comps["PR"], value=external)
                elif "CA" in comps:
                    # Theory
                    ws.cell(row=row, column=comps["CA"], value=internal)
                    
                    # Split external into MSE and ESE
                    if "MSE" in comps and "ESE" in comps:
                        # Find the subject to get max marks
                        curr_sub = next((s for s in subjects if s.subject_code == sub_code or 
                                        (s.is_elective and col_info == subject_col_map.get(s.subject_code))), None)
                        
                        if curr_sub and curr_sub.external_max > 0:
                            ext_max = curr_sub.external_max
                            if ext_max == 80:
                                mse_max, ese_max = 30, 50
                            else:
                                mse_max = round(ext_max * 0.375)
                                ese_max = ext_max - mse_max
                            
                            # Proportionally split
                            mse_marks = round(external * (mse_max / ext_max), 1) if ext_max > 0 else 0
                            ese_marks = round(external - mse_marks, 1)
                            
                            ws.cell(row=row, column=comps["MSE"], value=mse_marks)
                            ws.cell(row=row, column=comps["ESE"], value=ese_marks)
                
                # Fill TOT
                if "tot_col" in col_info:
                    tot_cell = ws.cell(row=row, column=col_info["tot_col"], value=total)
                    tot_cell.font = Font(bold=True)
            
            # Summary columns
            ws.cell(row=row, column=summary_start, value=total_marks)
            
            sgpa_cell = ws.cell(row=row, column=summary_start + 1, value=sem_record.sgpa)
            sgpa_cell.font = Font(bold=True, color="2F5496")
            
            result = "PASS" if sem_record.sgpa >= 4.0 else "FAIL"
            result_cell = ws.cell(row=row, column=summary_start + 2, value=result)
            result_cell.font = Font(bold=True, color="006600" if result == "PASS" else "CC0000")
        else:
            students_without_marks += 1
        
        # Apply borders to all columns
        for c in range(1, total_cols + 1):
            ws.cell(row=row, column=c).border = BORDER
            if c > 3:
                ws.cell(row=row, column=c).alignment = Alignment(horizontal="center")
    
    print(f"✅ Students with marks: {students_with_marks}")
    print(f"⏳ Students without marks: {students_without_marks}")
    
    # Freeze panes
    ws.freeze_panes = f"D{DATA_START}"
    
    # ── Subject Info Sheet ──
    ws2 = wb.create_sheet("Subject Info")
    info_headers = ["Code", "Name", "Credits", "Type", "Internal Max", "External Max", "Total Max"]
    
    for i, h in enumerate(info_headers, 1):
        cell = ws2.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="548235")
        cell.border = BORDER
    
    for ri, sub in enumerate(subjects, 2):
        vals = [
            sub.subject_code, sub.subject_name, sub.credits, sub.course_type,
            sub.internal_max, sub.external_max, sub.internal_max + sub.external_max
        ]
        for ci, v in enumerate(vals, 1):
            ws2.cell(row=ri, column=ci, value=v).border = BORDER
    
    for ci in range(1, len(info_headers) + 1):
        ws2.column_dimensions[get_column_letter(ci)].width = 25
    
    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


async def export_all_semesters():
    """Export marks for all semesters that have data"""
    
    client = await init_database()
    
    try:
        from app.models.student_profile import StudentProfile
        
        BRANCH = "IT"
        ADMISSION_YEAR = 2022
        ACADEMIC_YEAR = "2024-25"
        OUTPUT_DIR = "./exported_marks"
        
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        
        print(f"\n📊 Exporting all semester marks for {BRANCH} - {ADMISSION_YEAR}")
        
        # Find which semesters have marks
        students = await StudentProfile.find({
            "branch": {"$regex": f"^{BRANCH}$", "$options": "i"},
            "admission_year": ADMISSION_YEAR,
        }).to_list()
        
        semesters_with_marks = set()
        for student in students:
            for sr in student.semester_records:
                semesters_with_marks.add(sr.semester_number)
        
        if not semesters_with_marks:
            print("⚠️ No marks found for any semester!")
            return
        
        print(f"Found marks for semesters: {sorted(semesters_with_marks)}")
        
        for sem in sorted(semesters_with_marks):
            try:
                buf = await generate_marks_template_direct(
                    semester=sem,
                    branch=BRANCH,
                    academic_year=ACADEMIC_YEAR,
                    admission_year=ADMISSION_YEAR,
                )
                
                filename = f"marks_sem{sem}_{BRANCH}_{ACADEMIC_YEAR.replace('-', '_')}.xlsx"
                filepath = os.path.join(OUTPUT_DIR, filename)
                
                with open(filepath, 'wb') as f:
                    f.write(buf.read())
                
                print(f"  ✅ Semester {sem}: {filepath}")
                
            except Exception as e:
                print(f"  ❌ Semester {sem}: {e}")
        
        print(f"\n✅ Export complete! Files saved to: {os.path.abspath(OUTPUT_DIR)}")
        
    finally:
        client.close()


async def list_students_with_marks():
    """List all students and their marks status"""
    
    client = await init_database()
    
    try:
        from app.models.student_profile import StudentProfile
        
        BRANCH = "IT"
        ADMISSION_YEAR = 2022
        
        students = await StudentProfile.find({
            "branch": {"$regex": f"^{BRANCH}$", "$options": "i"},
            "admission_year": ADMISSION_YEAR,
        }).sort("roll_number").to_list()
        
        print(f"\n{'='*80}")
        print(f"📊 Students in {BRANCH} - Admission {ADMISSION_YEAR}")
        print(f"{'='*80}")
        print(f"Total: {len(students)}")
        print(f"{'='*80}\n")
        
        print(f"{'No.':<5} {'Roll Number':<15} {'Name':<25} {'CGPA':<8} {'Semesters with Marks'}")
        print("-" * 80)
        
        for i, s in enumerate(students, 1):
            sems = [str(sr.semester_number) for sr in s.semester_records]
            sems_str = ", ".join(sems) if sems else "None"
            print(f"{i:<5} {s.roll_number:<15} {s.name:<25} {s.cgpa:<8.2f} {sems_str}")
        
        print()
        
    finally:
        client.close()


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Export marks templates")
    parser.add_argument("--all", action="store_true", help="Export all semesters with marks")
    parser.add_argument("--list", action="store_true", help="List students with marks status")
    parser.add_argument("--semester", "-s", type=int, help="Specific semester to export")
    parser.add_argument("--branch", "-b", type=str, default="IT", help="Branch (default: IT)")
    parser.add_argument("--admission-year", "-a", type=int, default=2022, help="Admission year (default: 2022)")
    
    args = parser.parse_args()
    
    if args.list:
        asyncio.run(list_students_with_marks())
    elif args.all:
        asyncio.run(export_all_semesters())
    else:
        asyncio.run(export_marks())