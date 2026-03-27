"""
Fix existing Excel templates to support both roll numbers and seat numbers
"""

import sys
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def fix_template(input_file: str, output_file: str = None):
    """Fix an existing template to support both identifiers"""
    
    if output_file is None:
        output_file = input_file.replace('.xlsx', '_fixed.xlsx')
    
    wb = load_workbook(input_file)
    ws = wb["Marks Data"]
    
    # Update the Seat No header to be clearer
    seat_cell = ws.cell(row=4, column=2)
    seat_cell.value = "Roll No / Seat No\n(Primary ID)"
    
    # Insert a new column for alternate ID
    ws.insert_cols(4)
    
    # Add header for alternate ID
    alt_header = ws.cell(row=4, column=4, value="Alt. ID\n(Optional)")
    alt_header.font = Font(bold=True, color="FFFFFF", size=10)
    alt_header.fill = PatternFill("solid", fgColor="2F5496")
    alt_header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alt_header.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    ws.merge_cells(start_row=4, start_column=4, end_row=5, end_column=4)
    
    # Set column width
    ws.column_dimensions['D'].width = 12
    
    # Save
    wb.save(output_file)
    print(f"✅ Fixed template saved to: {output_file}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python fix_excel_headers.py <input_file.xlsx> [output_file.xlsx]")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    fix_template(input_file, output_file)